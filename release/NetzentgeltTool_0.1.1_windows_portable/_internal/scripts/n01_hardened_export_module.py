from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    TEMPLATE_DIR,
    NutzungsmeldungExportResult,
    _as_ru_tuple,
    _fetch_usage_segments,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
)
from ukl_preflight_module import (
    PreflightIssue,
    raise_if_blocking_issues,
    validate_n01_rows,
)
from zuordnungen_export_module import LTE_HOLDING_MARKET_PARTNER_IDS


N01_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Übernahmeanfrage,Übergabemeldung.xlsx"
N01_HEADERS = (
    "TfzE oder tEns*",
    "Beginn der Nutzung*",
    "Ende der Nutzung",
    "Nutzer-vEns*",
    "Marktpartner ID für Nutzungsüberlassung*",
)


def _validate_lte_holding_recipients(rows: list[dict[str, object]]) -> list[PreflightIssue]:
    """Im aktuellen LTE-Scope ausschließlich bestätigte Holding-Empfänger zulassen."""
    issues: list[PreflightIssue] = []

    for row_number, row in enumerate(rows, start=1):
        recipient = str(row.get("holder_market_partner_id") or "").strip()

        if recipient and recipient not in LTE_HOLDING_MARKET_PARTNER_IDS:
            issues.append(
                PreflightIssue(
                    code="N01_RECIPIENT_NOT_LTE_HOLDING",
                    message=(
                        "Empfänger-Marktpartner-ID ist nicht einer der beiden bestätigten "
                        "LTE-Holding-Mandanten zugeordnet."
                    ),
                    row_number=row_number,
                )
            )

    return issues


def build_hardened_n01_xlsx(
    *,
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = N01_TEMPLATE_PATH,
) -> NutzungsmeldungExportResult:
    """Aktuelle N01-Datei mit exakt fünf Spalten erzeugen."""
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Aktuelle UKL-N01-Vorlage fehlt: {template_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = _fetch_usage_segments(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        header_id, header_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    raise_if_blocking_issues(
        [
            *validate_n01_rows(rows),
            *_validate_lte_holding_recipients(rows),
        ],
        export_name="N01-Nutzungsmeldung",
    )

    workbook = load_workbook(template_path)
    worksheet = workbook["Zuordnungsdatensatzliste"]
    if worksheet.max_column > 5:
        worksheet.delete_cols(6, worksheet.max_column - 5)
    worksheet["A1"] = "Nutzung einer tEns"
    worksheet["B2"] = "N01"
    for column, header in enumerate(N01_HEADERS, start=1):
        worksheet.cell(row=6, column=column).value = header
    _prepare_template_rows(worksheet, required_data_rows=len(rows), max_column=5)
    worksheet["B3"] = str(header_id or "")
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_name

    for offset, row in enumerate(rows):
        target = 7 + offset
        values = (
            str(row["locomotive_no"]),
            row["usage_start"],
            row["usage_end"],
            str(row["user_vens"] or ""),
            str(row["holder_market_partner_id"] or ""),
        )
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=target, column=column).value = value
        for column in (1, 4, 5):
            worksheet.cell(row=target, column=column).number_format = "@"
        for column in (2, 3):
            worksheet.cell(row=target, column=column).number_format = "dd.mm.yyyy hh:mm"

    output = BytesIO()
    workbook.save(output)
    return NutzungsmeldungExportResult(
        content=output.getvalue(),
        file_name=(
            f"Nutzungsmeldung_{_safe_file_part(export_label)}_"
            f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
        ),
        row_count=len(rows),
        missing_required_mapping_count=0,
    )
