from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    AUFENTHALTSEREIGNIS_TEMPLATE_PATH,
    AufenthaltsereignisExportResult,
    _as_ru_tuple,
    _assert_export_gate_ready,
    _placeholders,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
)
from ukl_preflight_module import raise_if_blocking_issues, validate_ae01_rows


def _fetch_hardened_ae01_rows(
    con,
    *,
    performing_ru_values: tuple[str, ...],
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """AE01-Ereignisse mit echter gemappter vEns statt PerformingRU-Firmennamen lesen."""
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)

    rows = con.execute(
        f"""
        with movement_base as (
            select
                cast(loco_no as varchar) as locomotive_no,
                performing_ru,
                nullif(trim(cast(user_vens as varchar)), '') as user_vens,
                upper(coalesce(faulty_dir, '')) as faulty_dir_norm,
                upper(coalesce(clean_dir, '')) as clean_dir_norm,
                report_scope,
                sequence_ts,
                actual_departure_ts,
                actual_arrival_ts,
                origin_name,
                destination_name
            from core_loco_timeline
            where row_type = 'MOVEMENT'
              and nullif(trim(loco_no), '') is not null
              and performing_ru in ({placeholders})
              and coalesce(needs_manual_review, false) = false
        ),
        primary_events as (
            select
                locomotive_no,
                performing_ru,
                user_vens,
                case
                    when faulty_dir_norm = 'E' then destination_name
                    when faulty_dir_norm = 'A' then origin_name
                    when clean_dir_norm in ('E', 'E/A') then origin_name
                    when clean_dir_norm = 'A' then destination_name
                    else coalesce(origin_name, destination_name)
                end as event_location,
                case
                    when faulty_dir_norm = 'E' then actual_arrival_ts
                    when faulty_dir_norm = 'A' then actual_departure_ts
                    when clean_dir_norm in ('E', 'E/A') then actual_departure_ts
                    when clean_dir_norm = 'A' then actual_arrival_ts
                    else coalesce(sequence_ts, actual_departure_ts, actual_arrival_ts)
                end as event_ts,
                case
                    when faulty_dir_norm = 'E' then 'einfahrend'
                    when faulty_dir_norm = 'A' then 'ausfahrend'
                    when clean_dir_norm in ('E', 'E/A') then 'einfahrend'
                    when clean_dir_norm = 'A' then 'ausfahrend'
                    when report_scope = 'IN_REPORT' then 'netzintern'
                    else 'netzextern'
                end as network_status
            from movement_base
        ),
        clean_double_exit as (
            select
                locomotive_no,
                performing_ru,
                user_vens,
                destination_name as event_location,
                actual_arrival_ts as event_ts,
                'ausfahrend' as network_status
            from movement_base
            where clean_dir_norm = 'E/A'
              and faulty_dir_norm not in ('E', 'A')
        )
        select *
        from (
            select * from primary_events
            union all
            select * from clean_double_exit
        ) all_events
        where event_ts >= ?
          and event_ts < ?
        order by locomotive_no, event_ts, network_status
        """,
        [*ru_values, window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "performing_ru": row[1],
            "user_vens": row[2],
            "event_location": row[3],
            "event_ts": row[4],
            "network_status": row[5],
        }
        for row in rows
    ]


def build_hardened_aufenthaltsereignis_xlsx(
    *,
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = AUFENTHALTSEREIGNIS_TEMPLATE_PATH,
) -> AufenthaltsereignisExportResult:
    """AE01-Datei nur mit gemappter vEns und erfolgreichem UKL-Preflight erzeugen."""
    db_path = Path(db_path)
    template_path = Path(template_path)
    ru_values = _as_ru_tuple(performing_ru_values)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"Aktuelle UKL-AE01-Vorlage fehlt: {template_path}")

    con = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = _fetch_hardened_ae01_rows(
            con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        header_id, header_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    raise_if_blocking_issues(validate_ae01_rows(rows), export_name="AE01-Aufenthaltsereignis")

    workbook = load_workbook(template_path)
    worksheet = workbook["Aufenthaltsereignisse"]
    _prepare_template_rows(worksheet, required_data_rows=len(rows), first_data_row=8, max_column=5)
    worksheet["B3"] = str(header_id or "")
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_name

    for offset, row in enumerate(rows):
        target = 8 + offset
        values = (
            str(row["locomotive_no"]),
            str(row["user_vens"] or ""),
            str(row["event_location"] or ""),
            row["event_ts"],
            str(row["network_status"]),
        )
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=target, column=column).value = value
        for column in (1, 2, 5):
            worksheet.cell(row=target, column=column).number_format = "@"
        worksheet.cell(row=target, column=4).number_format = "dd.mm.yyyy hh:mm"

    output = BytesIO()
    workbook.save(output)
    return AufenthaltsereignisExportResult(
        content=output.getvalue(),
        file_name=(
            f"Aufenthaltsereignis_{_safe_file_part(export_label)}_"
            f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
        ),
        row_count=len(rows),
        missing_required_field_count=0,
    )
