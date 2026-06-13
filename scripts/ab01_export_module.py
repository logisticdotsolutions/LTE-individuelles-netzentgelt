from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    TEMPLATE_DIR,
    _as_ru_tuple,
    _assert_export_gate_ready,
    _placeholders,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
)


AB01_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Abstellungen.xlsx"

# Abstellungsart für Kaltabstellungen aus core_loco_stand_candidates.
# Offene UKL-Frage (Spec Abschnitt 23.7): Wie wird eine kalte Abstellung
# formal exportiert? Bis zur schriftlichen Klärung durch UKL: "TfzE nicht in Nutzung".
_ART_COLD_STAND = "TfzE nicht in Nutzung"


@dataclass(frozen=True)
class Ab01ExportResult:
    content: bytes
    file_name: str
    row_count: int
    missing_required_field_count: int


def _fetch_abstellungen(
    con,
    performing_ru_values: tuple[str, ...],
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """Potenzielle Kaltabstellungen aus core_loco_stand_candidates ermitteln.

    Quelle: Standzeiten > 480 min am selben DE-Ort (stand_class='POTENTIAL_COLD_STAND').
    UKL-Abstellungsart: 'TfzE nicht in Nutzung' – offene Frage gemäß Spec Abschnitt 23.7.
    """
    ru_values = performing_ru_values
    placeholders = _placeholders(ru_values)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)

    rows = con.execute(
        f"""
        select
            cast(loco_no as varchar) as locomotive_no,
            performing_ru,
            stand_from_utc,
            stand_to_utc,
            location_name
        from core_loco_stand_candidates
        where performing_ru in ({placeholders})
          and stand_from_utc >= ?
          and stand_from_utc < ?
        order by loco_no asc, stand_from_utc asc
        """,
        [*ru_values, window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "performing_ru": row[1],
            "stand_from": row[2],
            "stand_to": row[3],
            "location_name": row[4],
        }
        for row in rows
    ]


def build_ab01_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = AB01_TEMPLATE_PATH,
) -> Ab01ExportResult:
    """UKL-XLSX-Abstellungen je PerformingRU als Download-Bytes erzeugen.

    Spalten der UKL-Vorlage AB01:
    A: TfzE oder tEns* = LocomotiveNo
    B: vEns* = PerformingRU
    C: Art* = 'TfzE nicht in Nutzung' (offene UKL-Frage Spec Abschnitt 23.7)
    D: Beginn* = stand_from_utc
    E: Ende* = stand_to_utc

    Datenquelle: core_loco_stand_candidates (Standzeiten > 480 min am selben Ort).
    """
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"XLSX-Vorlage fehlt: {template_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_abstellungen(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    workbook = load_workbook(template_path)
    if "Abstellungen" not in workbook.sheetnames:
        raise RuntimeError(
            "Die XLSX-Vorlage enthält das erwartete Tabellenblatt "
            "'Abstellungen' nicht."
        )

    worksheet = workbook["Abstellungen"]
    _prepare_template_rows(
        worksheet,
        required_data_rows=len(rows),
        first_data_row=8,
        max_column=5,
    )

    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name or " / ".join(ru_values)

    first_data_row = 8
    for offset, row in enumerate(rows):
        target = first_data_row + offset

        worksheet.cell(row=target, column=1).value = str(row["locomotive_no"])
        worksheet.cell(row=target, column=1).number_format = "@"

        worksheet.cell(row=target, column=2).value = str(row["performing_ru"])
        worksheet.cell(row=target, column=2).number_format = "@"

        worksheet.cell(row=target, column=3).value = _ART_COLD_STAND
        worksheet.cell(row=target, column=3).number_format = "@"

        worksheet.cell(row=target, column=4).value = row["stand_from"]
        worksheet.cell(row=target, column=4).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=target, column=5).value = row["stand_to"]
        worksheet.cell(row=target, column=5).number_format = "dd.mm.yyyy hh:mm"

    missing_required_field_count = sum(
        1
        for r in rows
        if not r["locomotive_no"]
        or not r["performing_ru"]
        or not r["stand_from"]
        or not r["stand_to"]
    )

    output = BytesIO()
    workbook.save(output)

    return Ab01ExportResult(
        content=output.getvalue(),
        file_name=(
            f"Abstellungen_{_safe_file_part(export_label)}_"
            f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
        ),
        row_count=len(rows),
        missing_required_field_count=missing_required_field_count,
    )
