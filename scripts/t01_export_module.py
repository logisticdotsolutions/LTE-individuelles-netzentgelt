from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    TEMPLATE_DIR,
    _as_ru_tuple,
    _assert_export_gate_ready,
    _placeholders,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
)
from t01_mapping_module import (
    load_classification_mapping,
    load_locomotive_characteristics,
    resolve_classification,
    resolve_locomotive_characteristics,
)
from t01_preflight_module import validate_t01_rows
from ukl_preflight_module import raise_if_blocking_issues
from ukl_vens_mapping_module import apply_vens_mapping


T01_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Traktionsleistungen.xlsx"


@dataclass(frozen=True)
class T01ExportResult:
    content: bytes
    file_name: str
    row_count: int


def _qident(value: str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def _columns(con, table_name: str) -> list[str]:
    return [row[0] for row in con.execute(f"describe {_qident(table_name)}").fetchall()]


def _pick(columns: Iterable[str], candidates: Iterable[str]) -> str | None:
    by_lower = {str(column).lower(): str(column) for column in columns}
    for candidate in candidates:
        if str(candidate).lower() in by_lower:
            return by_lower[str(candidate).lower()]
    return None


def _text(column: str | None) -> str:
    return "null" if not column else f"nullif(trim(cast({_qident(column)} as varchar)), '')"


def _timestamp(column: str | None) -> str:
    return "null" if not column else f"try_cast({_text(column)} as timestamp)"


def _number(column: str | None) -> str:
    return "null" if not column else f"try_cast(replace({_text(column)}, ',', '.') as double)"


def _fetch_raw_rows(con, *, performing_ru_values: tuple[str, ...], date_from: date, date_to: date):
    ru_values = _as_ru_tuple(performing_ru_values)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)
    columns = _columns(con, "raw_locomotivemovement")
    placeholders = _placeholders(ru_values)
    window_start, window_end = _to_day_bounds(date_from, date_to)

    performing_ru = _pick(columns, ["PerformingRU"])
    loco_no = _pick(columns, ["LocomotiveNo"])
    departure = _pick(columns, ["ActualDeparture", "LocomotiveActualDeparture"])
    arrival = _pick(columns, ["ActualArrival", "LocomotiveActualArrival"])
    origin_code = _pick(columns, ["OriginLocationCode", "LocomotiveOriginLocationCode"])
    destination_code = _pick(columns, ["DestinationLocationCode", "LocomotiveDestinationLocationCode"])
    origin_iso = _pick(columns, ["OriginCountryISO", "OriginCountry"])
    destination_iso = _pick(columns, ["DestinationCountryISO", "DestinationCountry"])
    country = _pick(columns, ["Country"])
    distance = _pick(columns, ["CalculatedDistance", "Distance", "RealKm", "Km"])
    gross_weight = _pick(columns, ["TrainWeightGross"])
    traction_weight = _pick(columns, ["TractionLocomotiveWeight"])
    train_no = _pick(columns, ["TrainNo", "OriginTrainNo", "DestinationTrainNo"])
    transport_no = _pick(columns, ["TransportNumber", "TransportNumberRu"])
    movement_type = _pick(columns, ["MovementType"])
    transport_type = _pick(columns, ["TransportType"])
    traction_type = _pick(columns, ["TractionType"])
    train_type = _pick(columns, ["TrainTypeEN"])

    if not performing_ru or not loco_no or not departure:
        raise RuntimeError("T01-Export nicht möglich: Pflichtspalten PerformingRU, LocomotiveNo oder ActualDeparture fehlen.")

    de_filters = [
        f"upper(coalesce({_text(column)}, '')) = 'DE'"
        for column in (origin_iso, destination_iso, country)
        if column
    ]
    de_filter = " or ".join(de_filters) if de_filters else "false"
    trailer_weight = f"({_number(gross_weight)} - {_number(traction_weight)})"

    rows = con.execute(
        f"""
        select
            {_text(loco_no)} as locomotive_no,
            {_text(performing_ru)} as performing_ru,
            {_timestamp(departure)} as departure_ts,
            {_text(origin_code)} as departure_location,
            {_timestamp(arrival)} as arrival_ts,
            {_text(destination_code)} as arrival_location,
            {_number(distance)} as distance_km,
            {trailer_weight} as trailer_weight_t,
            {_text(train_no)} as train_no,
            {_text(transport_no)} as transport_number,
            {_text(movement_type)} as movement_type,
            {_text(transport_type)} as transport_type,
            {_text(traction_type)} as traction_type,
            {_text(train_type)} as train_type_en
        from raw_locomotivemovement
        where {_text(performing_ru)} in ({placeholders})
          and {_timestamp(departure)} >= ?
          and {_timestamp(departure)} < ?
          and ({de_filter})
        order by locomotive_no, departure_ts
        """,
        [*ru_values, window_start, window_end],
    ).fetchall()

    keys = (
        "locomotive_no", "performing_ru", "departure_ts", "departure_location",
        "arrival_ts", "arrival_location", "distance_km", "trailer_weight_t",
        "train_no", "transport_number", "movement_type", "transport_type",
        "traction_type", "train_type_en",
    )
    return [dict(zip(keys, row)) for row in rows]


def enrich_t01_rows(rows):
    classifications = load_classification_mapping()
    characteristics = load_locomotive_characteristics()
    prepared = apply_vens_mapping(rows, timestamp_keys=("departure_ts",))
    result = []
    for source in prepared:
        row = dict(source)
        order_criterion, usage_type = resolve_classification(row, classifications)
        loco = resolve_locomotive_characteristics(
            loco_no=row.get("locomotive_no"),
            at_utc=row.get("departure_ts"),
            mappings=characteristics,
        )
        row["order_criterion"] = order_criterion
        row["usage_type"] = usage_type
        row["max_speed_kmh"] = loco["max_speed_kmh"]
        row["is_multiple_unit"] = loco["is_multiple_unit"]
        departure = row.get("departure_ts")
        row["traffic_day"] = departure.date() if departure else None
        result.append(row)
    return result


def fetch_t01_rows(*, db_path: Path, performing_ru_values: Iterable[str], date_from: date, date_to: date):
    db_path = Path(db_path)
    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        return enrich_t01_rows(_fetch_raw_rows(
            con,
            performing_ru_values=_as_ru_tuple(performing_ru_values),
            date_from=date_from,
            date_to=date_to,
        ))
    finally:
        con.close()


def build_t01_xlsx(
    *,
    db_path: Path,
    performing_ru_values: Iterable[str],
    virtual_extraction_point: str,
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = T01_TEMPLATE_PATH,
) -> T01ExportResult:
    db_path = Path(db_path)
    template_path = Path(template_path)
    ru_values = _as_ru_tuple(performing_ru_values)
    if not template_path.exists():
        raise FileNotFoundError(f"Aktuelle UKL-T01-Vorlage fehlt: {template_path}")

    rows = [
        row for row in fetch_t01_rows(
            db_path=db_path,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        if str(row.get("user_vens") or "").strip() == str(virtual_extraction_point).strip()
    ]
    if not rows:
        raise RuntimeError("Keine T01-Zeilen für die gewählte virtuelle Entnahmestelle gefunden.")
    raise_if_blocking_issues(validate_t01_rows(rows), export_name="T01-Traktionsleistungen")

    con = duckdb.connect(str(db_path), read_only=True)
    try:
        header_id, header_name = _resolve_export_header(con=con, performing_ru_values=ru_values)
    finally:
        con.close()

    workbook = load_workbook(template_path)
    worksheet = workbook["Traktionsleistungen"]
    _prepare_template_rows(worksheet, required_data_rows=len(rows), first_data_row=9, max_column=12)
    worksheet["B3"] = str(header_id or "")
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_name
    worksheet["B5"] = str(virtual_extraction_point)
    worksheet["B5"].number_format = "@"

    for offset, row in enumerate(rows):
        target = 9 + offset
        values = (
            row.get("locomotive_no"), row.get("departure_ts"), row.get("departure_location"),
            row.get("arrival_ts"), row.get("arrival_location"), row.get("distance_km"),
            row.get("trailer_weight_t"), row.get("train_no"), row.get("order_criterion"),
            row.get("usage_type"), row.get("max_speed_kmh"), row.get("traffic_day"),
        )
        for column, value in enumerate(values, start=1):
            worksheet.cell(row=target, column=column).value = value
        for column in (1, 3, 5, 8, 9, 10):
            worksheet.cell(row=target, column=column).number_format = "@"
        for column in (2, 4):
            worksheet.cell(row=target, column=column).number_format = "dd.mm.yyyy hh:mm"
        worksheet.cell(row=target, column=12).number_format = "dd.mm.yyyy"

    output = BytesIO()
    workbook.save(output)
    return T01ExportResult(
        content=output.getvalue(),
        file_name=(
            f"Traktionsleistungen_{_safe_file_part(export_label)}_"
            f"{_safe_file_part(virtual_extraction_point)}_"
            f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
        ),
        row_count=len(rows),
    )
