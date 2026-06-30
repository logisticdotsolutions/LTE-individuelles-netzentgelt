from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    NUTZUNGSMELDUNG_TEMPLATE_PATH,
    TEMPLATE_DIR,
    _as_ru_tuple,
    _fetch_usage_segments,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
    table_exists,
)


ZUORDNUNGEN_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Zuordnungen.xlsx"
ZUORDNUNGEN_HEADERS = (
    "TfzE oder tEns*",
    "Beginn der Zuordnung*",
    "Ende der Zuordnung",
    "Nutzer-vEns*",
    "Marktpartner ID für Nutzungsüberlassung",
)

LTE_HOLDING_MARKET_PARTNER_NAME = "LTE Logistik- und Transport-GmbH"
LTE_HOLDING_RAILCUBE_NAME = "LTE Logistik- und Transport-GmbH (Holding)"
LTE_HOLDING_MARKET_PARTNER_IDS = (
    "1900100300393",
    "1900100400391",
)
LTE_HOLDING_HOLDER_NAMES = (
    LTE_HOLDING_MARKET_PARTNER_NAME,
    LTE_HOLDING_RAILCUBE_NAME,
)


@dataclass(frozen=True)
class ZuordnungenExportResult:
    """Ergebnis eines dynamisch erzeugten UKL-XLSX-Zuordnungs-Exports."""

    content: bytes
    file_name: str
    row_count: int
    missing_required_field_count: int


def _sql_literal_list(values: Iterable[str]) -> str:
    cleaned_values = [str(value).strip() for value in values if str(value).strip()]
    if not cleaned_values:
        return "''"
    return ", ".join("'" + value.replace("'", "''") + "'" for value in cleaned_values)


def _holding_holder_filter_sql(alias: str = "s") -> str:
    """SQL-Filter für Segmente, deren Halter fachlich LTE Holding ist."""
    return f"""
        (
            trim(coalesce(cast({alias}.holder_market_partner_id as varchar), ''))
                in ({_sql_literal_list(LTE_HOLDING_MARKET_PARTNER_IDS)})
            or lower(trim(coalesce(cast({alias}.holder_name as varchar), '')))
                in ({_sql_literal_list(name.lower() for name in LTE_HOLDING_HOLDER_NAMES)})
        )
    """


def _load_zuordnungen_workbook(template_path: Path):
    """Offizielle Zuordnungsvorlage laden und auf das definierte Schema härten."""
    requested_template_path = Path(template_path)
    effective_template_path = (
        requested_template_path
        if requested_template_path.exists()
        else NUTZUNGSMELDUNG_TEMPLATE_PATH
    )

    if not effective_template_path.exists():
        raise FileNotFoundError(
            "XLSX-Vorlage fehlt. Erwartet wurde entweder "
            f"{requested_template_path} oder die versionierte Fallback-Vorlage "
            f"{NUTZUNGSMELDUNG_TEMPLATE_PATH}."
        )

    workbook = load_workbook(effective_template_path)

    if "Zuordnungsdatensatzliste" not in workbook.sheetnames:
        raise RuntimeError(
            "Die XLSX-Vorlage enthält das erwartete Tabellenblatt "
            "'Zuordnungsdatensatzliste' nicht."
        )

    worksheet = workbook["Zuordnungsdatensatzliste"]

    if worksheet.max_column > len(ZUORDNUNGEN_HEADERS):
        worksheet.delete_cols(
            len(ZUORDNUNGEN_HEADERS) + 1,
            worksheet.max_column - len(ZUORDNUNGEN_HEADERS),
        )

    worksheet["A1"] = "Zuordnung"
    worksheet["B2"] = "Z01"

    for column_number, header in enumerate(ZUORDNUNGEN_HEADERS, start=1):
        worksheet.cell(row=6, column=column_number).value = header

    return workbook


def _assert_holding_export_gate_ready(
    con,
    date_from: date,
    date_to: date,
) -> None:
    """Holding-Export nur bei blockierten LTE-Holding-Haltersegmenten verhindern."""
    required_tables = [
        "core_usage_assignment_segments",
        "core_usage_assignment_segment_movements",
        "dq_export_gate_ru",
        "dq_global_export_blockers",
    ]
    missing_tables = [
        table_name
        for table_name in required_tables
        if not table_exists(con, table_name)
    ]

    if missing_tables:
        raise RuntimeError(
            "Export-Gate fehlt. Pipeline neu ausführen. Fehlende Tabellen: "
            + ", ".join(missing_tables)
        )

    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)

    local_blockers = con.execute(
        f"""
        select
            count(*) as blocker_count,
            string_agg(
                distinct cast(g.coverage_date as varchar) || ': ' || g.loco_no,
                ', '
            ) as examples
        from dq_export_gate_ru g
        where g.coverage_date >= ?
          and g.coverage_date <= ?
          and g.gate_status = 'BLOCKED'
          and exists (
                select 1
                from core_usage_assignment_segments s
                where s.loco_no = g.loco_no
                  and s.performing_ru is not distinct from g.performing_ru
                  and {_holding_holder_filter_sql('s')}
                  and exists (
                        select 1
                        from core_usage_assignment_segment_movements m
                        where m.usage_segment_id = s.usage_segment_id
                          and m.actual_departure_ts >= ?
                          and m.actual_departure_ts < ?
                  )
          )
        """,
        [date_from, date_to, window_start, window_end_exclusive],
    ).fetchone()

    global_blockers = con.execute(
        """
        select
            count(*) as blocker_count,
            string_agg(
                distinct cast(blocker_date as varchar) || ': ' || rule_id,
                ', '
            ) as examples
        from dq_global_export_blockers
        where blocker_date >= ?
          and blocker_date <= ?
          and gate_status = 'BLOCKED'
        """,
        [date_from, date_to],
    ).fetchone()

    local_count = int(local_blockers[0] or 0)
    global_count = int(global_blockers[0] or 0)

    if local_count > 0 or global_count > 0:
        details = []

        if local_count > 0:
            details.append(
                f"Blockierte LTE-Holding-Halter-Lok-Tage: {local_count}. "
                f"Beispiele: {local_blockers[1] or '-'}"
            )

        if global_count > 0:
            details.append(
                f"Globale Blocker im Zeitraum: {global_count}. "
                f"Beispiele: {global_blockers[1] or '-'}"
            )

        raise RuntimeError(
            "Holding-Zuordnungsexport ist gesperrt, bis die blockierenden "
            "Prüffälle geklärt sind. " + " | ".join(details)
        )


def _fetch_holding_assignment_segments(
    con,
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """Exportfähige Z01-Segmente nur für Halter = LTE Holding liefern."""
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    _assert_holding_export_gate_ready(con, date_from, date_to)

    rows = con.execute(
        f"""
        select
            cast(s.loco_no as varchar) as locomotive_no,
            s.segment_start_utc,
            s.segment_end_utc,
            s.performing_ru,
            s.movement_count,
            coalesce(nullif(s.user_vens, ''), s.performing_ru) as user_vens,
            coalesce(nullif(s.holder_market_partner_id, ''), s.holder_name) as holder_market_partner_id
        from core_usage_assignment_segments s
        where coalesce(s.export_blocking_movement_rows, 0) = 0
          and {_holding_holder_filter_sql('s')}
          and exists (
                select 1
                from core_usage_assignment_segment_movements m
                where m.usage_segment_id = s.usage_segment_id
                  and m.actual_departure_ts >= ?
                  and m.actual_departure_ts < ?
          )
        order by s.loco_no, s.segment_start_utc
        """,
        [window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "usage_start": row[1],
            "usage_end": row[2],
            "performing_ru": row[3],
            "movement_count": row[4],
            "user_vens": row[5],
            "holder_market_partner_id": row[6],
        }
        for row in rows
    ]


def _build_zuordnungen_workbook_result(
    *,
    rows: list[dict[str, object]],
    header_market_partner_id: str,
    header_market_partner_name: str,
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path,
) -> ZuordnungenExportResult:
    """Gemeinsamen XLSX-Schreibpfad für RU- und Holding-Exporte ausführen."""
    workbook = _load_zuordnungen_workbook(Path(template_path))
    worksheet = workbook["Zuordnungsdatensatzliste"]
    _prepare_template_rows(
        worksheet,
        required_data_rows=len(rows),
        first_data_row=7,
        max_column=5,
    )

    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name

    first_data_row = 7

    for offset, export_row in enumerate(rows):
        row_number = first_data_row + offset

        worksheet.cell(row=row_number, column=1).value = str(export_row["locomotive_no"])
        worksheet.cell(row=row_number, column=1).number_format = "@"

        worksheet.cell(row=row_number, column=2).value = export_row["usage_start"]
        worksheet.cell(row=row_number, column=2).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=row_number, column=3).value = export_row["usage_end"]
        worksheet.cell(row=row_number, column=3).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=row_number, column=4).value = (
            str(export_row["user_vens"])
            if export_row["user_vens"] is not None
            else ""
        )
        worksheet.cell(row=row_number, column=4).number_format = "@"

        worksheet.cell(row=row_number, column=5).value = (
            str(export_row["holder_market_partner_id"])
            if export_row["holder_market_partner_id"] is not None
            else ""
        )
        worksheet.cell(row=row_number, column=5).number_format = "@"

    missing_required_field_count = sum(
        1
        for row in rows
        if not row["locomotive_no"]
        or not row["usage_start"]
        or not row["user_vens"]
    )

    output = BytesIO()
    workbook.save(output)

    file_name = (
        "Zuordnungen_"
        f"{_safe_file_part(export_label)}_"
        f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
    )

    return ZuordnungenExportResult(
        content=output.getvalue(),
        file_name=file_name,
        row_count=len(rows),
        missing_required_field_count=missing_required_field_count,
    )


def build_zuordnungen_holding_xlsx(
    db_path: Path,
    holding_market_partner_id: str,
    date_from: date,
    date_to: date,
    template_path: Path = ZUORDNUNGEN_TEMPLATE_PATH,
) -> ZuordnungenExportResult:
    """Eine Z01-Datei für LTE-Holding-Haltersegmente erzeugen."""
    db_path = Path(db_path)
    holding_market_partner_id = str(holding_market_partner_id).strip()

    if holding_market_partner_id not in LTE_HOLDING_MARKET_PARTNER_IDS:
        raise ValueError(
            "Unbekannte LTE-Holding-Marktpartner-ID: "
            f"{holding_market_partner_id or '-'}"
        )

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_holding_assignment_segments(
            con=con,
            date_from=date_from,
            date_to=date_to,
        )
    finally:
        con.close()

    return _build_zuordnungen_workbook_result(
        rows=rows,
        header_market_partner_id=holding_market_partner_id,
        header_market_partner_name=LTE_HOLDING_MARKET_PARTNER_NAME,
        export_label=f"LTE_Holding_{holding_market_partner_id}",
        date_from=date_from,
        date_to=date_to,
        template_path=template_path,
    )


def build_zuordnungen_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = ZUORDNUNGEN_TEMPLATE_PATH,
) -> ZuordnungenExportResult:
    """Bestehenden RU-spezifischen Z01-Backendpfad für Kompatibilität erhalten."""
    db_path = Path(db_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_usage_segments(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    return _build_zuordnungen_workbook_result(
        rows=rows,
        header_market_partner_id=header_market_partner_id,
        header_market_partner_name=header_market_partner_name,
        export_label=export_label,
        date_from=date_from,
        date_to=date_to,
        template_path=template_path,
    )
