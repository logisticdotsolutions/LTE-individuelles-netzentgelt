from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable, Sequence

import duckdb
from openpyxl import load_workbook

from export_module import (
    AUFENTHALTSEREIGNIS_TEMPLATE_PATH,
    NUTZUNGSMELDUNG_TEMPLATE_PATH,
    AufenthaltsereignisExportResult,
    NutzungsmeldungExportResult,
    _as_ru_tuple,
    _assert_export_gate_ready,
    _placeholders,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
    table_exists,
)


HOLDER_GROUPED_EXPORT_RUNTIME_MARKER = "NETZENTGELT_HOLDER_GROUPED_EXPORT_PHASE14D_V1_20260630"
UNRESOLVED_HOLDER_KEY = "UNGEKLAERTER_HALTER"


@dataclass(frozen=True)
class HolderExportGroup:
    holder_key: str
    holder_label: str
    holder_market_partner_id: str
    holder_name: str
    row_count: int
    locomotive_count: int


def _holder_key_sql(alias: str = "s") -> str:
    return (
        f"coalesce(nullif(trim(cast({alias}.holder_market_partner_id as varchar)), ''), "
        f"nullif(trim(cast({alias}.holder_name as varchar)), ''), "
        f"'{UNRESOLVED_HOLDER_KEY}')"
    )


def _holder_name_sql(alias: str = "s") -> str:
    return (
        f"coalesce(nullif(trim(cast({alias}.holder_name as varchar)), ''), "
        f"nullif(trim(cast({alias}.holder_market_partner_id as varchar)), ''), "
        f"'{UNRESOLVED_HOLDER_KEY}')"
    )


def _holder_market_partner_sql(alias: str = "s") -> str:
    return f"coalesce(nullif(trim(cast({alias}.holder_market_partner_id as varchar)), ''), '')"


def _holder_display_label(
    holder_key: str,
    holder_name: str,
    holder_market_partner_id: str,
) -> str:
    holder_name = str(holder_name or "").strip()
    holder_market_partner_id = str(holder_market_partner_id or "").strip()
    holder_key = str(holder_key or "").strip() or UNRESOLVED_HOLDER_KEY

    if holder_key == UNRESOLVED_HOLDER_KEY:
        return "Halter offen / nicht zugeordnet"

    if holder_market_partner_id and holder_name and holder_market_partner_id != holder_name:
        return f"{holder_name} ({holder_market_partner_id})"

    return holder_name or holder_market_partner_id or holder_key


def _holder_file_suffix(
    holder_key: str,
    holder_name: str,
    holder_market_partner_id: str,
) -> str:
    label_source = holder_market_partner_id or holder_name or holder_key or UNRESOLVED_HOLDER_KEY
    return _safe_file_part(label_source)


def list_holder_export_groups(
    db_path: Path,
    performing_ru_values: Iterable[str],
    date_from: date,
    date_to: date,
) -> list[HolderExportGroup]:
    """Exportfähige Segmente einer RU-Gruppe zusätzlich nach Halter bündeln."""
    db_path = Path(db_path)
    if not db_path.exists():
        return []

    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    holder_key_expr = _holder_key_sql("s")
    holder_name_expr = _holder_name_sql("s")
    holder_market_partner_expr = _holder_market_partner_sql("s")

    con = duckdb.connect(str(db_path), read_only=True)
    try:
        if not table_exists(con, "core_usage_assignment_segments") or not table_exists(
            con,
            "core_usage_assignment_segment_movements",
        ):
            return []

        rows = con.execute(
            f"""
            select
                {holder_key_expr} as holder_key,
                {holder_name_expr} as holder_name,
                {holder_market_partner_expr} as holder_market_partner_id,
                count(*) as row_count,
                count(distinct cast(s.loco_no as varchar)) as locomotive_count
            from core_usage_assignment_segments s
            where s.performing_ru in ({placeholders})
              and coalesce(s.export_blocking_movement_rows, 0) = 0
              and exists (
                    select 1
                    from core_usage_assignment_segment_movements m
                    where m.usage_segment_id = s.usage_segment_id
                      and m.actual_departure_ts >= ?
                      and m.actual_departure_ts < ?
              )
            group by 1, 2, 3
            order by holder_name, holder_market_partner_id, holder_key
            """,
            [*ru_values, window_start, window_end_exclusive],
        ).fetchall()
    finally:
        con.close()

    return [
        HolderExportGroup(
            holder_key=str(row[0]),
            holder_label=_holder_display_label(row[0], row[1], row[2]),
            holder_name=str(row[1] or ""),
            holder_market_partner_id=str(row[2] or ""),
            row_count=int(row[3] or 0),
            locomotive_count=int(row[4] or 0),
        )
        for row in rows
    ]


def _fetch_usage_segments_for_holder(
    con,
    performing_ru_values: Sequence[str],
    date_from: date,
    date_to: date,
    holder_key: str,
) -> list[dict[str, object]]:
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)

    if not table_exists(con, "core_usage_assignment_segments"):
        raise RuntimeError("core_usage_assignment_segments fehlt. Phase 6C Pipeline neu ausführen.")

    rows = con.execute(
        f"""
        select
            cast(s.loco_no as varchar) as locomotive_no,
            s.segment_start_utc,
            s.segment_end_utc,
            s.performing_ru,
            s.movement_count,
            coalesce(nullif(s.user_vens, ''), s.performing_ru) as user_vens,
            coalesce(nullif(s.holder_market_partner_id, ''), s.holder_name) as holder_market_partner_id,
            {_holder_key_sql('s')} as holder_key
        from core_usage_assignment_segments s
        where s.performing_ru in ({placeholders})
          and coalesce(s.export_blocking_movement_rows, 0) = 0
          and {_holder_key_sql('s')} = ?
          and exists (
                select 1
                from core_usage_assignment_segment_movements m
                where m.usage_segment_id = s.usage_segment_id
                  and m.actual_departure_ts >= ?
                  and m.actual_departure_ts < ?
          )
        order by s.loco_no, s.segment_start_utc
        """,
        [*ru_values, holder_key, window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "usage_start": row[1],
            "usage_end": row[2],
            "performing_ru": row[3],
            "movement_count": row[4],
            "user_vens": row[5],
            "holder_market_partner_id": row[6],
            "holder_key": row[7],
        }
        for row in rows
    ]


def _fetch_aufenthaltsereignisse_for_holder(
    con,
    performing_ru_values: Sequence[str],
    date_from: date,
    date_to: date,
    holder_key: str,
) -> list[dict[str, object]]:
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)

    rows = con.execute(
        f"""
        with movement_base as (
            select
                cast(loco_no as varchar) as locomotive_no,
                performing_ru,
                upper(coalesce(faulty_dir, '')) as faulty_dir_norm,
                upper(coalesce(clean_dir, '')) as clean_dir_norm,
                report_scope,
                sequence_ts,
                actual_departure_ts,
                actual_arrival_ts,
                origin_name,
                destination_name
            from core_loco_timeline
            where row_type = 'MOVEMENT'
              and nullif(trim(loco_no), '') is not null
              and performing_ru in ({placeholders})
              and coalesce(needs_manual_review, false) = false
        ),
        primary_events as (
            select
                locomotive_no,
                performing_ru,
                case
                    when faulty_dir_norm = 'E' then destination_name
                    when faulty_dir_norm = 'A' then origin_name
                    when clean_dir_norm in ('E', 'E/A') then origin_name
                    when clean_dir_norm = 'A' then destination_name
                    else coalesce(origin_name, destination_name)
                end as event_location,
                case
                    when faulty_dir_norm = 'E' then actual_arrival_ts
                    when faulty_dir_norm = 'A' then actual_departure_ts
                    when clean_dir_norm in ('E', 'E/A') then actual_departure_ts
                    when clean_dir_norm = 'A' then actual_arrival_ts
                    else coalesce(sequence_ts, actual_departure_ts, actual_arrival_ts)
                end as event_ts,
                case
                    when faulty_dir_norm = 'E' then 'einfahrend'
                    when faulty_dir_norm = 'A' then 'ausfahrend'
                    when clean_dir_norm in ('E', 'E/A') then 'einfahrend'
                    when clean_dir_norm = 'A' then 'ausfahrend'
                    when report_scope = 'IN_REPORT' then 'netzintern'
                    else 'netzextern'
                end as network_status
            from movement_base
        ),
        clean_double_exit as (
            select
                locomotive_no,
                performing_ru,
                destination_name as event_location,
                actual_arrival_ts as event_ts,
                'ausfahrend' as network_status
            from movement_base
            where clean_dir_norm = 'E/A'
              and faulty_dir_norm not in ('E', 'A')
        ),
        all_events as (
            select * from primary_events
            union all
            select * from clean_double_exit
        ),
        matched_events as (
            select
                e.locomotive_no,
                e.performing_ru,
                e.event_location,
                e.event_ts,
                e.network_status,
                row_number() over (
                    partition by e.locomotive_no, e.performing_ru, e.event_location, e.event_ts, e.network_status
                    order by s.segment_start_utc desc
                ) as match_rank
            from all_events e
            join core_usage_assignment_segments s
              on cast(s.loco_no as varchar) = e.locomotive_no
             and s.performing_ru is not distinct from e.performing_ru
             and e.event_ts >= s.segment_start_utc
             and e.event_ts <= s.segment_end_utc
            where e.event_ts >= ?
              and e.event_ts < ?
              and coalesce(s.export_blocking_movement_rows, 0) = 0
              and {_holder_key_sql('s')} = ?
        )
        select
            locomotive_no,
            performing_ru,
            event_location,
            event_ts,
            network_status
        from matched_events
        where match_rank = 1
        order by locomotive_no asc, event_ts asc, network_status asc
        """,
        [*ru_values, window_start, window_end_exclusive, holder_key],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "performing_ru": row[1],
            "event_location": row[2],
            "event_ts": row[3],
            "network_status": row[4],
        }
        for row in rows
    ]


def build_nutzungsmeldung_holder_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    holder_key: str,
    holder_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = NUTZUNGSMELDUNG_TEMPLATE_PATH,
) -> NutzungsmeldungExportResult:
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"XLSX-Vorlage fehlt: {template_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = _fetch_usage_segments_for_holder(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
            holder_key=holder_key,
        )
        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    workbook = load_workbook(template_path)
    if "Zuordnungsdatensatzliste" not in workbook.sheetnames:
        raise RuntimeError("Die XLSX-Vorlage enthält das Tabellenblatt 'Zuordnungsdatensatzliste' nicht.")

    worksheet = workbook["Zuordnungsdatensatzliste"]
    _prepare_template_rows(worksheet, required_data_rows=len(rows))
    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name

    first_data_row = 7
    for offset, export_row in enumerate(rows):
        row_number = first_data_row + offset
        worksheet.cell(row=row_number, column=1).value = str(export_row["locomotive_no"])
        worksheet.cell(row=row_number, column=1).number_format = "@"
        worksheet.cell(row=row_number, column=2).value = export_row["usage_start"]
        worksheet.cell(row=row_number, column=2).number_format = "dd.mm.yyyy hh:mm"
        worksheet.cell(row=row_number, column=3).value = export_row["usage_end"]
        worksheet.cell(row=row_number, column=3).number_format = "dd.mm.yyyy hh:mm"
        worksheet.cell(row=row_number, column=4).value = str(export_row["user_vens"] or "")
        worksheet.cell(row=row_number, column=4).number_format = "@"
        worksheet.cell(row=row_number, column=5).value = str(export_row["holder_market_partner_id"] or "")
        worksheet.cell(row=row_number, column=5).number_format = "@"
        worksheet.cell(row=row_number, column=6).value = None

    missing_required_mapping_count = sum(
        1
        for row in rows
        if not row["user_vens"] or not row["holder_market_partner_id"]
    )

    output = BytesIO()
    workbook.save(output)

    holder_suffix = _safe_file_part(holder_label or holder_key)
    file_name = (
        "Nutzungsmeldung_"
        f"{_safe_file_part(export_label)}_Halter_{holder_suffix}_"
        f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
    )

    return NutzungsmeldungExportResult(
        content=output.getvalue(),
        file_name=file_name,
        row_count=len(rows),
        missing_required_mapping_count=missing_required_mapping_count,
    )


def build_aufenthaltsereignis_holder_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    holder_key: str,
    holder_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = AUFENTHALTSEREIGNIS_TEMPLATE_PATH,
) -> AufenthaltsereignisExportResult:
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")
    if not template_path.exists():
        raise FileNotFoundError(f"XLSX-Vorlage fehlt: {template_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)
    try:
        rows = _fetch_aufenthaltsereignisse_for_holder(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
            holder_key=holder_key,
        )
        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    workbook = load_workbook(template_path)
    if "Aufenthaltsereignisse" not in workbook.sheetnames:
        raise RuntimeError("Die XLSX-Vorlage enthält das Tabellenblatt 'Aufenthaltsereignisse' nicht.")

    worksheet = workbook["Aufenthaltsereignisse"]
    _prepare_template_rows(worksheet, required_data_rows=len(rows), first_data_row=8, max_column=5)
    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name or " / ".join(ru_values)

    first_data_row = 8
    for offset, export_row in enumerate(rows):
        row_number = first_data_row + offset
        worksheet.cell(row=row_number, column=1).value = str(export_row["locomotive_no"])
        worksheet.cell(row=row_number, column=1).number_format = "@"
        worksheet.cell(row=row_number, column=2).value = str(export_row["performing_ru"])
        worksheet.cell(row=row_number, column=2).number_format = "@"
        worksheet.cell(row=row_number, column=3).value = str(export_row["event_location"] or "")
        worksheet.cell(row=row_number, column=4).value = export_row["event_ts"]
        worksheet.cell(row=row_number, column=4).number_format = "dd.mm.yyyy hh:mm"
        worksheet.cell(row=row_number, column=5).value = str(export_row["network_status"])
        worksheet.cell(row=row_number, column=5).number_format = "@"

    missing_required_field_count = sum(
        1
        for row in rows
        if not row["locomotive_no"]
        or not row["performing_ru"]
        or not row["event_location"]
        or not row["event_ts"]
        or not row["network_status"]
    )

    output = BytesIO()
    workbook.save(output)

    holder_suffix = _safe_file_part(holder_label or holder_key)
    file_name = (
        "Aufenthaltsereignis_"
        f"{_safe_file_part(export_label)}_Halter_{holder_suffix}_"
        f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
    )

    return AufenthaltsereignisExportResult(
        content=output.getvalue(),
        file_name=file_name,
        row_count=len(rows),
        missing_required_field_count=missing_required_field_count,
    )


def _render_holder_grouped_primary_download_card(
    *,
    group_key: str,
    group_config: dict,
    export_kind: str,
    db_path: Path,
    export_date_from: date,
    export_date_to: date,
    findings,
    export_gate_ru,
    global_export_blockers,
    build_nutzungsmeldung_download_cached,
    build_aufenthaltsereignis_download_cached,
) -> None:
    import streamlit as st
    import export_cockpit_ui_module as cockpit

    try:
        holder_groups = list_holder_export_groups(
            db_path=db_path,
            performing_ru_values=tuple(group_config["performing_ru_values"]),
            date_from=export_date_from,
            date_to=export_date_to,
        )
    except Exception as error:
        cockpit._render_open_case_overview(
            group_config=group_config,
            context_label="Halteraufteilung konnte noch nicht vorbereitet werden",
            findings=findings,
            export_gate_ru=export_gate_ru,
            global_export_blockers=global_export_blockers,
            technical_error=error,
        )
        return

    if not holder_groups:
        st.info("Für diese LTE-Gruppe wurden im Zeitraum keine exportfähigen Halter-Segmente gefunden.")
        return

    st.caption(f"{len(holder_groups)} Halter im gewählten Zeitraum")

    for holder_group in holder_groups:
        expander_label = (
            f"{holder_group.holder_label} · "
            f"{holder_group.row_count} Segmente · "
            f"{holder_group.locomotive_count} Loks"
        )
        with st.expander(expander_label, expanded=len(holder_groups) == 1):
            try:
                if export_kind == "nutzung":
                    result = build_nutzungsmeldung_holder_xlsx(
                        db_path=db_path,
                        performing_ru_values=tuple(group_config["performing_ru_values"]),
                        export_label=group_config["file_label"],
                        holder_key=holder_group.holder_key,
                        holder_label=holder_group.holder_label,
                        date_from=export_date_from,
                        date_to=export_date_to,
                    )
                    missing_count = result.missing_required_mapping_count
                    button_label = "Nutzung XLSX"
                else:
                    result = build_aufenthaltsereignis_holder_xlsx(
                        db_path=db_path,
                        performing_ru_values=tuple(group_config["performing_ru_values"]),
                        export_label=group_config["file_label"],
                        holder_key=holder_group.holder_key,
                        holder_label=holder_group.holder_label,
                        date_from=export_date_from,
                        date_to=export_date_to,
                    )
                    missing_count = result.missing_required_field_count
                    button_label = "Aufenthalt XLSX"
            except Exception as error:
                cockpit._render_open_case_overview(
                    group_config=group_config,
                    context_label=f"{holder_group.holder_label}: Export konnte noch nicht vorbereitet werden",
                    findings=findings,
                    export_gate_ru=export_gate_ru,
                    global_export_blockers=global_export_blockers,
                    technical_error=error,
                )
                continue

            st.metric("Zeilen", result.row_count)
            if missing_count > 0:
                st.info(
                    "Hier ist noch etwas offen: "
                    f"{missing_count} Zeilen haben noch fehlende Pflichtfelder. "
                    "Bitte Fall prüfen."
                )

            st.download_button(
                label=button_label,
                data=result.content,
                file_name=result.file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=(
                    "download_holder_grouped_"
                    f"{export_kind}_{group_key.lower()}_"
                    f"{_safe_file_part(holder_group.holder_key)}"
                ),
                use_container_width=True,
            )


def install_holder_grouped_export_runtime() -> None:
    """Exportcockpit so patchen, dass LTE-Arbeitsdateien zusätzlich je Halter erscheinen."""
    import export_cockpit_ui_module as cockpit

    if getattr(cockpit, "_holder_grouped_export_runtime_installed", False):
        return

    cockpit._original_render_primary_download_card_holder_grouped = (  # type: ignore[attr-defined]
        cockpit._render_primary_download_card
    )
    cockpit._render_primary_download_card = _render_holder_grouped_primary_download_card
    cockpit._holder_grouped_export_runtime_installed = True
