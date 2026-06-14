from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Iterable

import duckdb
from openpyxl import load_workbook

from export_module import (
    TEMPLATE_DIR,
    _as_ru_tuple,
    _assert_export_gate_ready,
    _placeholders,
    _prepare_template_rows,
    _resolve_export_header,
    _safe_file_part,
    _to_day_bounds,
)


AV01_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Aufenthaltsabschnitt.xlsx"


@dataclass(frozen=True)
class Av01ExportResult:
    content: bytes
    file_name: str
    row_count: int
    missing_required_field_count: int


def _fetch_aufenthaltsabschnitte(
    con,
    performing_ru_values: tuple[str, ...],
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """DE-relevante Aufenthaltsabschnitte aus der Lokomotiv-Timeline ermitteln.

    Je MOVEMENT-Zeile wird ein Abschnitt mit Beginn/Beginn Ort, Ende/Ende Ort
    und Netzstatus erzeugt. Netzstatus-Ableitung analog zur AE01-Logik:
    faulty_dir hat Vorrang vor clean_dir; E/A-Bewegungen werden als 'einfahrend'
    klassifiziert (primäre Richtung Einfahrt).
    """
    ru_values = performing_ru_values
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    _assert_export_gate_ready(con, ru_values, date_from, date_to)

    rows = con.execute(
        f"""
        select
            cast(loco_no as varchar) as locomotive_no,
            performing_ru,
            actual_departure_ts,
            origin_name,
            actual_arrival_ts,
            destination_name,
            case
                when upper(coalesce(faulty_dir, '')) = 'E' then 'einfahrend'
                when upper(coalesce(faulty_dir, '')) = 'A' then 'ausfahrend'
                when upper(coalesce(clean_dir, '')) in ('E', 'E/A') then 'einfahrend'
                when upper(coalesce(clean_dir, '')) = 'A' then 'ausfahrend'
                when report_scope = 'IN_REPORT' then 'netzintern'
                else 'netzextern'
            end as network_status
        from core_loco_timeline
        where row_type = 'MOVEMENT'
          and performing_ru in ({placeholders})
          and coalesce(needs_manual_review, false) = false
          and actual_departure_ts >= ?
          and actual_departure_ts < ?
          and nullif(trim(loco_no), '') is not null
        order by loco_no asc, actual_departure_ts asc
        """,
        [*ru_values, window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "performing_ru": row[1],
            "departure_ts": row[2],
            "departure_location": row[3],
            "arrival_ts": row[4],
            "arrival_location": row[5],
            "network_status": row[6],
        }
        for row in rows
    ]


def build_av01_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = AV01_TEMPLATE_PATH,
) -> Av01ExportResult:
    """UKL-XLSX-Aufenthaltsabschnitte je PerformingRU als Download-Bytes erzeugen.

    Spalten der UKL-Vorlage AV01:
    A: TfzE oder tEns* = LocomotiveNo
    B: vEns* = PerformingRU
    C: Beginn* = ActualDeparture
    D: Beginn Ort = Origin
    E: Ende* = ActualArrival
    F: Ende Ort = Destination
    G: Netzstatus* = einfahrend / ausfahrend / netzintern / netzextern
    """
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    if not template_path.exists():
        raise FileNotFoundError(f"XLSX-Vorlage fehlt: {template_path}")

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_aufenthaltsabschnitte(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )
        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )
    finally:
        con.close()

    workbook = load_workbook(template_path)
    if "Aufenthaltsabschnitt" not in workbook.sheetnames:
        raise RuntimeError(
            "Die XLSX-Vorlage enthält das erwartete Tabellenblatt "
            "'Aufenthaltsabschnitt' nicht."
        )

    worksheet = workbook["Aufenthaltsabschnitt"]
    _prepare_template_rows(
        worksheet,
        required_data_rows=len(rows),
        first_data_row=8,
        max_column=7,
    )

    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name or " / ".join(ru_values)

    first_data_row = 8
    for offset, row in enumerate(rows):
        target = first_data_row + offset

        worksheet.cell(row=target, column=1).value = str(row["locomotive_no"])
        worksheet.cell(row=target, column=1).number_format = "@"

        worksheet.cell(row=target, column=2).value = str(row["performing_ru"])
        worksheet.cell(row=target, column=2).number_format = "@"

        worksheet.cell(row=target, column=3).value = row["departure_ts"]
        worksheet.cell(row=target, column=3).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=target, column=4).value = (
            str(row["departure_location"]) if row["departure_location"] else ""
        )

        worksheet.cell(row=target, column=5).value = row["arrival_ts"]
        worksheet.cell(row=target, column=5).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=target, column=6).value = (
            str(row["arrival_location"]) if row["arrival_location"] else ""
        )

        worksheet.cell(row=target, column=7).value = str(row["network_status"])
        worksheet.cell(row=target, column=7).number_format = "@"

    missing_required_field_count = sum(
        1
        for r in rows
        if not r["locomotive_no"]
        or not r["performing_ru"]
        or not r["departure_ts"]
        or not r["arrival_ts"]
        or not r["network_status"]
    )

    output = BytesIO()
    workbook.save(output)

    return Av01ExportResult(
        content=output.getvalue(),
        file_name=(
            f"Aufenthaltsabschnitt_{_safe_file_part(export_label)}_"
            f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
        ),
        row_count=len(rows),
        missing_required_field_count=missing_required_field_count,
    )
