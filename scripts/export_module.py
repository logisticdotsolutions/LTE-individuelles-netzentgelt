"""
Netzentgelt MVP - Exportmodul
============================

Zweck
-----
Dieses Modul bündelt sämtliche Exportaufgaben der Netzentgelt-Pipeline:

1. Aufbau der fachlichen CSV-Exporttabellen in DuckDB
2. Schreiben der fachlichen sowie Audit-/Debug-CSV-Dateien
3. Erzeugung der UKL-XLSX-Nutzungsmeldung je PerformingRU
4. Erzeugung der UKL-XLSX-Aufenthaltsereignisse je PerformingRU

Die XLSX-Ausgabe basiert auf der offiziellen Vorlage
``data/05_templates/Vorlage_Nutzungsmeldung.xlsx``.

Fachliche Segmentlogik der XLSX-Nutzungsmeldung
-----------------------------------------------
- Eine Exportzeile entspricht einer ununterbrochenen Nutzung einer Lok durch
  dieselbe PerformingRU.
- Eine GAP-Zeile beendet die laufende Nutzung.
- Ein Wechsel der PerformingRU beendet die laufende Nutzung ebenfalls.
- Nach einer GAP-Zeile kann dieselbe Lok erneut als weitere Exportzeile
  vorkommen.
- Der Datumsfilter entscheidet anhand von ActualDeparture, ob ein Segment in
  den Export aufgenommen wird. Der Filter gilt inklusive des gesamten
  ausgewählten Bis-Tages.
- Beginn und Ende bleiben die tatsächlichen Grenzen des vollständigen,
  ununterbrochenen Segments. Sie werden nicht auf den UI-Filter abgeschnitten.
- Sortierung: LocomotiveNo, danach Beginn der Nutzung.
"""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Iterable, Sequence
import re

import duckdb
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
EXP_DIR = ROOT / "data" / "03_exports"
TEMPLATE_DIR = ROOT / "data" / "05_templates"
NUTZUNGSMELDUNG_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Nutzungsmeldung.xlsx"
AUFENTHALTSEREIGNIS_TEMPLATE_PATH = TEMPLATE_DIR / "Vorlage_Aufenthaltsereignis.xlsx"


# ---------------------------------------------------------------------------
# Konfiguration der fixen LTE-Abschnitte in der Streamlit-Oberfläche.
# Die Werte entsprechen den PerformingRU-Schreibweisen aus RailCube/DataLake.
# Weitere Schreibweisen können hier ohne Änderung der Exportlogik ergänzt
# werden.
# ---------------------------------------------------------------------------
LTE_EXPORT_GROUPS = {
    "LTE_DE": {
        "title": "Performing RU LTE DE",
        "file_label": "LTE_DE",
        "performing_ru_values": (
            "LTE DE - LTE Germany GmbH",
            "LTE Germany GmbH",
        ),
    },
    "LTE_NL": {
        "title": "Performing RU LTE NL",
        "file_label": "LTE_NL",
        "performing_ru_values": (
            "LTE NL - LTE Netherlands B.V.",
        ),
    },
    "LTE_AT": {
        "title": "Performing RU LTE AT",
        "file_label": "LTE_AT",
        "performing_ru_values": (
            "LTE AT - LTE Austria GmbH",
        ),
    },
    "LTE_CH": {
        "title": "Performing RU LTE CH",
        "file_label": "LTE_CH",
        "performing_ru_values": (
            "LTE CH - LTE Schweiz GmbH",
        ),
    },
}


# Fachliche Exporte: Diese Dateien sind für die nachgelagerte Verarbeitung
# beziehungsweise die spätere Befüllung der UKL-Vorlagen relevant.
FACHLICHE_CSV_EXPORTS = [
    ("export_zuordnungen", "export_zuordnungen.csv"),
    ("export_nutzungsmeldung", "export_nutzungsmeldung.csv"),
]


# Audit- und Debug-Exporte: Diese Dateien dienen der Nachvollziehbarkeit,
# Fehleranalyse und fachlichen Freigabe.
AUDIT_CSV_EXPORTS = [
    ("raw_import_run", "raw_import_run.csv"),
    ("stg_loco_events", "stg_loco_events.csv"),
    ("core_loco_timeline", "core_loco_timeline.csv"),
    ("dq_findings", "dq_findings.csv"),
    ("cfg_dq_rule_catalog", "cfg_dq_rule_catalog.csv"),
    ("cfg_market_partner_role", "cfg_market_partner_role.csv"),
    ("cfg_market_partner_role_conflicts", "cfg_market_partner_role_conflicts.csv"),
    ("cfg_market_partner_mapping", "cfg_market_partner_mapping.csv"),
    ("cfg_market_partner_mapping_effective", "cfg_market_partner_mapping_effective.csv"),
    ("cfg_market_partner_mapping_conflicts", "cfg_market_partner_mapping_conflicts.csv"),
    ("cfg_market_partner_mapping_invalid", "cfg_market_partner_mapping_invalid.csv"),
    ("cfg_vens_tens_exception", "cfg_vens_tens_exception.csv"),
    ("cfg_vens_tens_exception_effective", "cfg_vens_tens_exception_effective.csv"),
    ("cfg_vens_tens_exception_conflicts", "cfg_vens_tens_exception_conflicts.csv"),
    (
        "dq_unresolved_performing_ru_market_partner_alias",
        "dq_unresolved_performing_ru_market_partner_alias.csv",
    ),
    ("stg_loco_events_skipped", "stg_loco_events_skipped.csv"),
    ("stg_transport_details_enriched", "stg_transport_details_enriched.csv"),
    ("core_transport_route", "core_transport_route.csv"),
]


ALL_CSV_EXPORTS = AUDIT_CSV_EXPORTS + FACHLICHE_CSV_EXPORTS


@dataclass(frozen=True)
class NutzungsmeldungExportResult:
    """Ergebnis eines dynamisch erzeugten UKL-XLSX-Exports."""

    content: bytes
    file_name: str
    row_count: int
    missing_required_mapping_count: int


@dataclass(frozen=True)
class AufenthaltsereignisExportResult:
    """Ergebnis eines dynamisch erzeugten UKL-XLSX-Aufenthaltsereignis-Exports."""

    content: bytes
    file_name: str
    row_count: int
    missing_required_field_count: int


def qident(name: str) -> str:
    """SQL-Identifier sicher quoten, beispielsweise DuckDB-Tabellennamen."""
    return '"' + name.replace('"', '""') + '"'


def table_exists(con, table_name: str) -> bool:
    """Prüfen, ob eine erwartete DuckDB-Tabelle existiert."""
    return (
        con.execute(
            """
            select count(*)
            from information_schema.tables
            where lower(table_name) = lower(?)
            """,
            [table_name],
        ).fetchone()[0]
        > 0
    )


def build_export_tables(con) -> None:
    """
    Bestehende fachliche CSV-Exporttabellen in DuckDB neu aufbauen.

    Diese Tabellen bleiben für Audit und Rückwärtskompatibilität erhalten.
    Der neue RU-bezogene XLSX-Export wird dynamisch über
    ``build_nutzungsmeldung_xlsx()`` erzeugt.
    """
    con.execute(
        """
        create or replace table export_zuordnungen as
        select
            tfze_or_tens as "TfzE oder tEns*",
            period_start_utc as "Beginn der Zuordnung*",
            period_end_utc as "Ende der Zuordnung",
            user_vens as "Nutzer-vEns*",
            performing_ru_marktpartner_id as "Marktpartner ID für Nutzungsüberlassung"
        from core_loco_timeline
        where row_type = 'MOVEMENT'
          and report_scope = 'IN_REPORT'
          and export_ready = true
        """
    )

    con.execute(
        """
        create or replace table export_nutzungsmeldung as
        select
            tfze_or_tens as "TfzE oder tEns*",
            period_start_utc as "Beginn der Nutzung*",
            period_end_utc as "Ende der Nutzung",
            coalesce(nullif(user_vens, ''), performing_ru) as "Nutzer-vEns*",
            coalesce(nullif(holder_market_partner_id, ''), holder_name) as "Marktpartner ID für Nutzungsüberlassung*",
            '' as "Übernahmeanfrage oder Übergabemeldung?"
        from core_loco_timeline
        where row_type = 'MOVEMENT'
          and report_scope = 'IN_REPORT'
          and export_ready = true
        """
    )


def export_table_to_csv(
    con,
    table_name: str,
    file_name: str,
    export_dir: Path = EXP_DIR,
) -> Path:
    """Eine DuckDB-Tabelle als semikolongetrennte CSV-Datei schreiben."""
    if not table_exists(con, table_name):
        raise RuntimeError(
            f"Export nicht möglich: Erwartete DuckDB-Tabelle fehlt: {table_name}"
        )

    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / file_name

    con.execute(
        f"copy {qident(table_name)} to ? (header true, delimiter ';')",
        [str(output_path)],
    )

    print(f"Export: {output_path}")
    return output_path


def export_all_csvs(con, export_dir: Path = EXP_DIR) -> list[Path]:
    """Alle fachlichen sowie Audit-/Debug-CSV-Dateien neu schreiben."""
    exported_paths = []

    for table_name, file_name in ALL_CSV_EXPORTS:
        exported_paths.append(
            export_table_to_csv(
                con=con,
                table_name=table_name,
                file_name=file_name,
                export_dir=export_dir,
            )
        )

    print(
        "CSV-Export abgeschlossen: "
        f"{len(exported_paths)} Dateien nach {export_dir} geschrieben."
    )

    return exported_paths


# ---------------------------------------------------------------------------
# Dynamischer XLSX-Export der Nutzungsmeldung
# ---------------------------------------------------------------------------


def _normalize_company_name_sql(expression: str) -> str:
    """Konservative Firmennamen-Normalisierung als SQL-Ausdruck liefern."""
    return f"""
        regexp_replace(
            lower(
                replace(
                    replace(
                        replace(
                            replace(
                                coalesce(cast({expression} as varchar), ''),
                                'ä', 'ae'
                            ),
                            'ö', 'oe'
                        ),
                        'ü', 'ue'
                    ),
                    'ß', 'ss'
                )
            ),
            '[^a-z0-9]+',
            '',
            'g'
        )
    """


def _as_ru_tuple(performing_ru_values: Iterable[str]) -> tuple[str, ...]:
    """PerformingRU-Werte trimmen, leere Werte entfernen und deduplizieren."""
    result = []

    for value in performing_ru_values:
        cleaned = str(value).strip()
        if cleaned and cleaned not in result:
            result.append(cleaned)

    return tuple(result)


def _placeholders(values: Sequence[object]) -> str:
    """SQL-Platzhalterliste passend zur Parameteranzahl bilden."""
    if not values:
        raise ValueError("Mindestens eine PerformingRU muss angegeben werden.")

    return ", ".join("?" for _ in values)


def _to_day_bounds(date_from: date, date_to: date) -> tuple[datetime, datetime]:
    """Inklusiven Datumsfilter in [Beginn, Folgetag) umwandeln."""
    if date_from > date_to:
        raise ValueError("Das Von-Datum darf nicht nach dem Bis-Datum liegen.")

    return (
        datetime.combine(date_from, time.min),
        datetime.combine(date_to + timedelta(days=1), time.min),
    )


def list_non_lte_performing_rus(db_path: Path) -> list[str]:
    """
    Alle DE-relevanten PerformingRUs außerhalb der vier fixen LTE-Gruppen liefern.

    Dadurch kann die Streamlit-Oberfläche im Abschnitt "Performing RU nicht LTE"
    eine konkrete RU auswählen. Der Export bleibt auch dort immer RU-spezifisch.
    """
    if not Path(db_path).exists():
        return []

    known_lte_values = sorted(
        {
            ru_value
            for config in LTE_EXPORT_GROUPS.values()
            for ru_value in config["performing_ru_values"]
        }
    )

    con = duckdb.connect(str(db_path), read_only=True)

    try:
        sql = """
            select distinct trim(performing_ru) as performing_ru
            from core_loco_timeline
            where row_type = 'MOVEMENT'
              and report_scope = 'IN_REPORT'
              and nullif(trim(performing_ru), '') is not null
        """
        params: list[object] = []

        if known_lte_values:
            sql += f" and trim(performing_ru) not in ({_placeholders(known_lte_values)})"
            params.extend(known_lte_values)

        # Nicht konfigurierte LTE-Schreibweisen dürfen nicht versehentlich
        # im Abschnitt "Performing RU nicht LTE" landen.
        sql += " and upper(trim(performing_ru)) not like 'LTE %'"
        sql += " order by performing_ru"

        return [row[0] for row in con.execute(sql, params).fetchall()]

    finally:
        con.close()


def list_unconfigured_lte_performing_rus(db_path: Path) -> list[str]:
    """LTE-Schreibweisen liefern, die noch keiner fixen LTE-Gruppe zugeordnet sind."""
    if not Path(db_path).exists():
        return []

    known_lte_values = sorted(
        {
            ru_value
            for config in LTE_EXPORT_GROUPS.values()
            for ru_value in config["performing_ru_values"]
        }
    )

    con = duckdb.connect(str(db_path), read_only=True)

    try:
        sql = """
            select distinct trim(performing_ru) as performing_ru
            from core_loco_timeline
            where row_type = 'MOVEMENT'
              and report_scope = 'IN_REPORT'
              and nullif(trim(performing_ru), '') is not null
              and upper(trim(performing_ru)) like 'LTE %'
        """
        params: list[object] = []

        if known_lte_values:
            sql += f" and trim(performing_ru) not in ({_placeholders(known_lte_values)})"
            params.extend(known_lte_values)

        sql += " order by performing_ru"

        return [row[0] for row in con.execute(sql, params).fetchall()]

    finally:
        con.close()


def _fetch_usage_segments(
    con,
    performing_ru_values: Sequence[str],
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """
    Ununterbrochene Nutzungssegmente für die gewählte PerformingRU ermitteln.

    Der Datumsfilter greift anhand der ActualDeparture-Werte der DE-relevanten
    Movement-Zeilen. Für ein Segment genügt mindestens eine passende Bewegung
    innerhalb des gewählten Tagesfensters.
    """
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)
    normalized_performing_ru_sql = _normalize_company_name_sql("s.performing_ru")

    rows = con.execute(
        f"""
        with ordered as (
            select
                c.*,

                lag(row_type) over (
                    partition by loco_no
                    order by
                        sort_sequence asc,
                        case when row_type = 'MOVEMENT' then 0 else 1 end,
                        source_row_id asc
                ) as previous_row_type,

                lag(performing_ru) over (
                    partition by loco_no
                    order by
                        sort_sequence asc,
                        case when row_type = 'MOVEMENT' then 0 else 1 end,
                        source_row_id asc
                ) as previous_performing_ru

            from core_loco_timeline c
            where nullif(trim(loco_no), '') is not null
        ),
        marked as (
            select
                *,
                case
                    when row_type <> 'MOVEMENT'
                        then 0

                    when previous_row_type is null
                        then 1

                    when previous_row_type = 'GAP'
                        then 1

                    when previous_performing_ru is distinct from performing_ru
                        then 1

                    else 0
                end as starts_new_usage_segment
            from ordered
        ),
        segmented as (
            select
                *,
                sum(starts_new_usage_segment) over (
                    partition by loco_no
                    order by
                        sort_sequence asc,
                        case when row_type = 'MOVEMENT' then 0 else 1 end,
                        source_row_id asc
                    rows between unbounded preceding and current row
                ) as usage_segment_no
            from marked
        ),
        segment_summary as (
            select
                loco_no,
                performing_ru,
                usage_segment_no,

                first(
                    actual_departure_ts
                    order by sort_sequence asc, source_row_id asc
                ) filter (where row_type = 'MOVEMENT') as usage_start,

                first(
                    actual_arrival_ts
                    order by sort_sequence desc, source_row_id desc
                ) filter (where row_type = 'MOVEMENT') as usage_end,

                count(*) filter (
                    where row_type = 'MOVEMENT'
                ) as movement_count,

                first(
                    nullif(trim(holder_name), '')
                    order by sort_sequence asc, source_row_id asc
                ) filter (
                    where row_type = 'MOVEMENT'
                      and nullif(trim(holder_name), '') is not null
                ) as holder_name,

                max(
                    case
                        when row_type = 'MOVEMENT'
                         and report_scope = 'IN_REPORT'
                         and actual_departure_ts >= ?
                         and actual_departure_ts < ?
                            then 1
                        else 0
                    end
                ) as matches_selected_departure_day

            from segmented
            group by
                loco_no,
                performing_ru,
                usage_segment_no
        )
        select
            cast(s.loco_no as varchar) as locomotive_no,
            s.usage_start,
            s.usage_end,
            s.performing_ru,
            s.movement_count,

            coalesce(
                anu_mapping.market_partner_id,
                anu_direct.market_partner_id,
                s.performing_ru
            ) as user_vens,

            coalesce(
                ane_mapping.market_partner_id,
                ane_direct.market_partner_id,
                s.holder_name
            ) as holder_market_partner_id

        from segment_summary s

        left join cfg_market_partner_mapping_effective anu_mapping
          on anu_mapping.role_code = 'ANU_VENS'
         and anu_mapping.source_value_normalized = {normalized_performing_ru_sql}

        left join cfg_market_partner_role_effective anu_direct
          on anu_direct.role_code = 'ANU_VENS'
         and anu_direct.company_name_normalized = {normalized_performing_ru_sql}

        left join cfg_market_partner_mapping_effective ane_mapping
          on ane_mapping.role_code = 'ANE_TENS'
         and ane_mapping.source_value_normalized = {normalized_performing_ru_sql}

        left join cfg_market_partner_role_effective ane_direct
          on ane_direct.role_code = 'ANE_TENS'
         and ane_direct.company_name_normalized = {normalized_performing_ru_sql}

        where s.performing_ru in ({placeholders})
          and s.matches_selected_departure_day = 1
          and s.usage_start is not null

        order by
            s.loco_no asc,
            s.usage_start asc
        """,
        [window_start, window_end_exclusive, *ru_values],
    ).fetchall()

    result = []

    for row in rows:
        result.append(
            {
                "locomotive_no": row[0],
                "usage_start": row[1],
                "usage_end": row[2],
                "performing_ru": row[3],
                "movement_count": row[4],
                "user_vens": row[5],
                "holder_market_partner_id": row[6],
            }
        )

    return result


def _resolve_export_header(
    con,
    performing_ru_values: Sequence[str],
) -> tuple[str, str]:
    """
    Marktpartner-Kopfdaten der Vorlage für eine RU-Gruppe ermitteln.

    Mehrere DataLake-Schreibweisen sind zulässig, sofern sie auf dieselbe
    eindeutige ANU_VENS-MP-ID zeigen, beispielsweise LTE DE.
    """
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    normalized_source_ru_sql = _normalize_company_name_sql("source_ru.performing_ru")

    rows = con.execute(
        f"""
        select distinct
            coalesce(
                anu_mapping.market_partner_id,
                anu_direct.market_partner_id
            ) as market_partner_id,

            coalesce(
                anu_mapping.official_company_name,
                anu_direct.company_name_official,
                source_ru.performing_ru
            ) as market_partner_name

        from (
            select unnest(?) as performing_ru
        ) source_ru

        left join cfg_market_partner_mapping_effective anu_mapping
          on anu_mapping.role_code = 'ANU_VENS'
         and anu_mapping.source_value_normalized = {normalized_source_ru_sql}

        left join cfg_market_partner_role_effective anu_direct
          on anu_direct.role_code = 'ANU_VENS'
         and anu_direct.company_name_normalized = {normalized_source_ru_sql}

        where source_ru.performing_ru in ({placeholders})
          and coalesce(
                anu_mapping.market_partner_id,
                anu_direct.market_partner_id
              ) is not null
        """,
        [list(ru_values), *ru_values],
    ).fetchall()

    unique_ids = sorted({str(row[0]).strip() for row in rows if row[0]})
    unique_names = sorted({str(row[1]).strip() for row in rows if row[1]})

    if len(unique_ids) > 1:
        raise RuntimeError(
            "Mehrere ANU_VENS-Marktpartner-IDs für dieselbe Exportgruppe gefunden: "
            + ", ".join(unique_ids)
        )

    market_partner_id = unique_ids[0] if unique_ids else ""
    market_partner_name = unique_names[0] if len(unique_names) == 1 else " / ".join(unique_names)

    return market_partner_id, market_partner_name


def _safe_file_part(value: str) -> str:
    """Text für einen Windows-tauglichen Dateinamen aufbereiten."""
    cleaned = re.sub(r"[^A-Za-z0-9_-]+", "_", str(value).strip())
    return cleaned.strip("_") or "PerformingRU"


def _copy_row_style(ws, source_row: int, target_row: int, max_column: int = 6) -> None:
    """Formatierung einer Vorlagenzeile auf eine neue Datenzeile übertragen."""
    for column in range(1, max_column + 1):
        source_cell = ws.cell(row=source_row, column=column)
        target_cell = ws.cell(row=target_row, column=column)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)


def _prepare_template_rows(
    ws,
    required_data_rows: int,
    first_data_row: int = 7,
    max_column: int = 6,
) -> None:
    """Bestehende Datenzeilen leeren und bei Bedarf formatiert erweitern."""
    template_style_row = (
        first_data_row + 1
        if ws.max_row >= first_data_row + 1
        else first_data_row
    )
    last_required_row = max(
        first_data_row,
        first_data_row + required_data_rows - 1,
    )

    if last_required_row > ws.max_row:
        for row_number in range(ws.max_row + 1, last_required_row + 1):
            _copy_row_style(
                ws=ws,
                source_row=template_style_row,
                target_row=row_number,
                max_column=max_column,
            )

    for row_number in range(first_data_row, max(ws.max_row, last_required_row) + 1):
        for column in range(1, max_column + 1):
            ws.cell(row=row_number, column=column).value = None


def build_nutzungsmeldung_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = NUTZUNGSMELDUNG_TEMPLATE_PATH,
) -> NutzungsmeldungExportResult:
    """
    UKL-XLSX-Nutzungsmeldung je PerformingRU als Download-Bytes erzeugen.

    Spalten der UKL-Vorlage:
    A: LocomotiveNo
    B: Beginn der ersten Bewegung der ununterbrochenen Nutzung
    C: Ende der letzten Bewegung der ununterbrochenen Nutzung
    D: ANU_VENS-MP-ID der PerformingRU; Fallback: PerformingRU
    E: ANE_TENS-MP-ID der PerformingRU; Fallback: Halter der Lok
    F: leer
    """
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    if not template_path.exists():
        raise FileNotFoundError(
            "XLSX-Vorlage fehlt. Erwartete Ablage: "
            f"{template_path}"
        )

    ru_values = _as_ru_tuple(performing_ru_values)

    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_usage_segments(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )

        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )

    finally:
        con.close()

    workbook = load_workbook(template_path)

    if "Zuordnungsdatensatzliste" not in workbook.sheetnames:
        raise RuntimeError(
            "Die XLSX-Vorlage enthält das erwartete Tabellenblatt "
            "'Zuordnungsdatensatzliste' nicht."
        )

    worksheet = workbook["Zuordnungsdatensatzliste"]
    _prepare_template_rows(worksheet, required_data_rows=len(rows))

    # Kopfdaten der Vorlage. Marktpartner-ID bewusst als Text schreiben.
    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name

    first_data_row = 7

    for offset, export_row in enumerate(rows):
        row_number = first_data_row + offset

        worksheet.cell(row=row_number, column=1).value = str(export_row["locomotive_no"])
        worksheet.cell(row=row_number, column=1).number_format = "@"

        worksheet.cell(row=row_number, column=2).value = export_row["usage_start"]
        worksheet.cell(row=row_number, column=2).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=row_number, column=3).value = export_row["usage_end"]
        worksheet.cell(row=row_number, column=3).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=row_number, column=4).value = (
            str(export_row["user_vens"])
            if export_row["user_vens"] is not None
            else ""
        )
        worksheet.cell(row=row_number, column=4).number_format = "@"

        worksheet.cell(row=row_number, column=5).value = (
            str(export_row["holder_market_partner_id"])
            if export_row["holder_market_partner_id"] is not None
            else ""
        )
        worksheet.cell(row=row_number, column=5).number_format = "@"

        worksheet.cell(row=row_number, column=6).value = None

    missing_required_mapping_count = sum(
        1
        for row in rows
        if not row["user_vens"] or not row["holder_market_partner_id"]
    )

    output = BytesIO()
    workbook.save(output)

    file_name = (
        "Nutzungsmeldung_"
        f"{_safe_file_part(export_label)}_"
        f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
    )

    return NutzungsmeldungExportResult(
        content=output.getvalue(),
        file_name=file_name,
        row_count=len(rows),
        missing_required_mapping_count=missing_required_mapping_count,
    )

# ---------------------------------------------------------------------------
# Dynamischer XLSX-Export der Aufenthaltsereignisse
# ---------------------------------------------------------------------------


def _fetch_aufenthaltsereignisse(
    con,
    performing_ru_values: Sequence[str],
    date_from: date,
    date_to: date,
) -> list[dict[str, object]]:
    """
    Aufenthaltsereignisse für die gewählte PerformingRU ermitteln.

    Ableitung je Movement-Zeile:
    - FaultyDir=E:  einfahrend, ActualArrival, Destination
    - FaultyDir=A:  ausfahrend, ActualDeparture, Origin
    - CleanDir=E:   einfahrend, ActualDeparture, Origin
    - CleanDir=A:   ausfahrend, ActualArrival, Destination
    - CleanDir=E/A: zwei Ereignisse: Einfahrt und Ausfahrt
    - sonstige DE-Zeile: netzintern
    - sonstige Nicht-DE-Zeile: netzextern

    Bei netzintern/netzextern wird der erste verfügbare Ort der Movement-Zeile
    verwendet. Der Datumsfilter greift tagesscharf auf den Ereigniszeitpunkt.
    """
    ru_values = _as_ru_tuple(performing_ru_values)
    placeholders = _placeholders(ru_values)
    window_start, window_end_exclusive = _to_day_bounds(date_from, date_to)

    rows = con.execute(
        f"""
        with movement_base as (
            select
                cast(loco_no as varchar) as locomotive_no,
                performing_ru,
                upper(coalesce(faulty_dir, '')) as faulty_dir_norm,
                upper(coalesce(clean_dir, '')) as clean_dir_norm,
                report_scope,
                sequence_ts,
                actual_departure_ts,
                actual_arrival_ts,
                origin_name,
                destination_name
            from core_loco_timeline
            where row_type = 'MOVEMENT'
              and nullif(trim(loco_no), '') is not null
              and performing_ru in ({placeholders})
        ),
        primary_events as (
            select
                locomotive_no,
                performing_ru,
                case
                    when faulty_dir_norm = 'E' then destination_name
                    when faulty_dir_norm = 'A' then origin_name
                    when clean_dir_norm in ('E', 'E/A') then origin_name
                    when clean_dir_norm = 'A' then destination_name
                    else coalesce(origin_name, destination_name)
                end as event_location,
                case
                    when faulty_dir_norm = 'E' then actual_arrival_ts
                    when faulty_dir_norm = 'A' then actual_departure_ts
                    when clean_dir_norm in ('E', 'E/A') then actual_departure_ts
                    when clean_dir_norm = 'A' then actual_arrival_ts
                    else coalesce(sequence_ts, actual_departure_ts, actual_arrival_ts)
                end as event_ts,
                case
                    when faulty_dir_norm = 'E' then 'einfahrend'
                    when faulty_dir_norm = 'A' then 'ausfahrend'
                    when clean_dir_norm in ('E', 'E/A') then 'einfahrend'
                    when clean_dir_norm = 'A' then 'ausfahrend'
                    when report_scope = 'IN_REPORT' then 'netzintern'
                    else 'netzextern'
                end as network_status
            from movement_base
        ),
        clean_double_exit as (
            select
                locomotive_no,
                performing_ru,
                destination_name as event_location,
                actual_arrival_ts as event_ts,
                'ausfahrend' as network_status
            from movement_base
            where clean_dir_norm = 'E/A'
              and faulty_dir_norm not in ('E', 'A')
        ),
        all_events as (
            select * from primary_events
            union all
            select * from clean_double_exit
        )
        select
            locomotive_no,
            performing_ru,
            event_location,
            event_ts,
            network_status
        from all_events
        where event_ts >= ?
          and event_ts < ?
        order by
            locomotive_no asc,
            event_ts asc,
            network_status asc
        """,
        [*ru_values, window_start, window_end_exclusive],
    ).fetchall()

    return [
        {
            "locomotive_no": row[0],
            "performing_ru": row[1],
            "event_location": row[2],
            "event_ts": row[3],
            "network_status": row[4],
        }
        for row in rows
    ]


def build_aufenthaltsereignis_xlsx(
    db_path: Path,
    performing_ru_values: Iterable[str],
    export_label: str,
    date_from: date,
    date_to: date,
    template_path: Path = AUFENTHALTSEREIGNIS_TEMPLATE_PATH,
) -> AufenthaltsereignisExportResult:
    """
    UKL-XLSX-Aufenthaltsereignisse je PerformingRU als Download-Bytes erzeugen.

    Spalten der UKL-Vorlage:
    A: TfzE oder tEns = LocomotiveNo
    B: vEns = PerformingRU
    C: Ort = Border Point beziehungsweise Movement-Ort
    D: Zeitpunkt = Border Time beziehungsweise Movement-Zeitpunkt
    E: Netzstatus = einfahrend / ausfahrend / netzintern / netzextern
    """
    db_path = Path(db_path)
    template_path = Path(template_path)

    if not db_path.exists():
        raise FileNotFoundError(f"DuckDB-Datei fehlt: {db_path}")

    if not template_path.exists():
        raise FileNotFoundError(
            "XLSX-Vorlage fehlt. Erwartete Ablage: "
            f"{template_path}"
        )

    ru_values = _as_ru_tuple(performing_ru_values)
    con = duckdb.connect(str(db_path), read_only=True)

    try:
        rows = _fetch_aufenthaltsereignisse(
            con=con,
            performing_ru_values=ru_values,
            date_from=date_from,
            date_to=date_to,
        )

        header_market_partner_id, header_market_partner_name = _resolve_export_header(
            con=con,
            performing_ru_values=ru_values,
        )

    finally:
        con.close()

    workbook = load_workbook(template_path)

    if "Aufenthaltsereignisse" not in workbook.sheetnames:
        raise RuntimeError(
            "Die XLSX-Vorlage enthält das erwartete Tabellenblatt "
            "'Aufenthaltsereignisse' nicht."
        )

    worksheet = workbook["Aufenthaltsereignisse"]
    _prepare_template_rows(
        worksheet,
        required_data_rows=len(rows),
        first_data_row=8,
        max_column=5,
    )

    worksheet["B3"] = str(header_market_partner_id) if header_market_partner_id else ""
    worksheet["B3"].number_format = "@"
    worksheet["B4"] = header_market_partner_name or " / ".join(ru_values)

    first_data_row = 8

    for offset, export_row in enumerate(rows):
        row_number = first_data_row + offset

        worksheet.cell(row=row_number, column=1).value = str(export_row["locomotive_no"])
        worksheet.cell(row=row_number, column=1).number_format = "@"

        worksheet.cell(row=row_number, column=2).value = str(export_row["performing_ru"])
        worksheet.cell(row=row_number, column=2).number_format = "@"

        worksheet.cell(row=row_number, column=3).value = (
            str(export_row["event_location"])
            if export_row["event_location"] is not None
            else ""
        )

        worksheet.cell(row=row_number, column=4).value = export_row["event_ts"]
        worksheet.cell(row=row_number, column=4).number_format = "dd.mm.yyyy hh:mm"

        worksheet.cell(row=row_number, column=5).value = str(export_row["network_status"])
        worksheet.cell(row=row_number, column=5).number_format = "@"

    missing_required_field_count = sum(
        1
        for row in rows
        if not row["locomotive_no"]
        or not row["performing_ru"]
        or not row["event_location"]
        or not row["event_ts"]
        or not row["network_status"]
    )

    output = BytesIO()
    workbook.save(output)

    file_name = (
        "Aufenthaltsereignis_"
        f"{_safe_file_part(export_label)}_"
        f"{date_from.isoformat()}_bis_{date_to.isoformat()}.xlsx"
    )

    return AufenthaltsereignisExportResult(
        content=output.getvalue(),
        file_name=file_name,
        row_count=len(rows),
        missing_required_field_count=missing_required_field_count,
    )

