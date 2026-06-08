from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path

import duckdb
import pytest
from openpyxl import Workbook, load_workbook

import export_module


def create_export_db(path: Path) -> None:
    con = duckdb.connect(str(path))
    try:
        con.execute("""
            create table core_usage_assignment_segments (
                usage_segment_id varchar, loco_no varchar, tfze_or_tens varchar,
                segment_start_utc timestamp, segment_end_utc timestamp,
                performing_ru varchar, movement_count bigint, user_vens varchar,
                holder_market_partner_id varchar, holder_name varchar,
                export_blocking_movement_rows bigint
            )
        """)
        con.execute("insert into core_usage_assignment_segments values ('SEG-1','91800000001-1','91800000001-1','2026-06-01 10:00:00','2026-06-01 11:00:00','RU GmbH',1,'VENS-RU','MP-HOLDER','Holder GmbH',0)")
        con.execute("create table core_usage_assignment_segment_movements (usage_segment_id varchar, actual_departure_ts timestamp)")
        con.execute("insert into core_usage_assignment_segment_movements values ('SEG-1','2026-06-01 10:00:00')")
        con.execute("create table dq_export_gate_ru (loco_no varchar, performing_ru varchar, coverage_date date, gate_status varchar)")
        con.execute("insert into dq_export_gate_ru values ('91800000001-1','RU GmbH','2026-06-01','READY')")
        con.execute("create table dq_global_export_blockers (blocker_date date, gate_status varchar)")
        con.execute("""
            create table cfg_market_partner_mapping_effective (
                role_code varchar, source_value_normalized varchar, source_value varchar,
                official_company_name varchar, market_partner_id varchar,
                match_method varchar, match_score double
            )
        """)
        con.execute("""
            create table cfg_market_partner_role_effective (
                role_code varchar, company_name_normalized varchar,
                company_name_official varchar, market_partner_id varchar
            )
        """)
        con.execute("insert into cfg_market_partner_role_effective values ('ANU_VENS','rugmbh','RU GmbH','MP-RU')")
        con.execute("""
            create table core_loco_timeline (
                row_type varchar, loco_no varchar, performing_ru varchar,
                faulty_dir varchar, clean_dir varchar, report_scope varchar,
                sequence_ts timestamp, actual_departure_ts timestamp,
                actual_arrival_ts timestamp, origin_name varchar,
                destination_name varchar, needs_manual_review boolean
            )
        """)
        con.execute("insert into core_loco_timeline values ('MOVEMENT','91800000001-1','RU GmbH',null,'E','IN_REPORT','2026-06-01 10:00:00','2026-06-01 10:00:00','2026-06-01 11:00:00','Border A','Border B',false)")
    finally:
        con.close()


def create_template(path: Path, sheet_name: str) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    for row in range(1, 10):
        for column in range(1, 7):
            worksheet.cell(row=row, column=column).value = None
    workbook.save(path)


@pytest.mark.integration
def test_semicolon_csv_export(con, tmp_path: Path):
    con.execute("create table export_fixture (A varchar, B bigint)")
    con.execute("insert into export_fixture values ('x', 1)")
    output = export_module.export_table_to_csv(con, "export_fixture", "fixture.csv", tmp_path)
    assert output.read_text(encoding="utf-8").splitlines() == ['A;B', 'x;1']


@pytest.mark.integration
def test_build_nutzungsmeldung_xlsx_from_temporary_db_and_template(tmp_path: Path):
    db_path = tmp_path / "export.duckdb"
    template = tmp_path / "Vorlage_Nutzungsmeldung.xlsx"
    create_export_db(db_path)
    create_template(template, "Zuordnungsdatensatzliste")
    result = export_module.build_nutzungsmeldung_xlsx(db_path, ["RU GmbH"], "RU", date(2026, 6, 1), date(2026, 6, 1), template)
    assert result.row_count == 1
    assert result.missing_required_mapping_count == 0
    workbook = load_workbook(BytesIO(result.content))
    ws = workbook["Zuordnungsdatensatzliste"]
    assert ws["B3"].value == "MP-RU"
    assert ws.cell(row=7, column=1).value == "91800000001-1"
    assert ws.cell(row=7, column=4).value == "VENS-RU"
    assert ws.cell(row=7, column=5).value == "MP-HOLDER"
    assert ws.cell(row=7, column=6).value is None


@pytest.mark.integration
def test_build_aufenthaltsereignis_xlsx_from_temporary_db_and_template(tmp_path: Path):
    db_path = tmp_path / "event.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsereignis.xlsx"
    create_export_db(db_path)
    create_template(template, "Aufenthaltsereignisse")
    result = export_module.build_aufenthaltsereignis_xlsx(db_path, ["RU GmbH"], "RU", date(2026, 6, 1), date(2026, 6, 1), template)
    assert result.row_count == 1
    assert result.missing_required_field_count == 0
    workbook = load_workbook(BytesIO(result.content))
    ws = workbook["Aufenthaltsereignisse"]
    assert ws["B3"].value == "MP-RU"
    assert ws.cell(row=8, column=1).value == "91800000001-1"
    assert ws.cell(row=8, column=2).value == "RU GmbH"
    assert ws.cell(row=8, column=3).value == "Border A"
    assert ws.cell(row=8, column=5).value == "einfahrend"


@pytest.mark.integration
def test_xlsx_export_refuses_blocked_gate(tmp_path: Path):
    db_path = tmp_path / "blocked.duckdb"
    template = tmp_path / "Vorlage_Nutzungsmeldung.xlsx"
    create_export_db(db_path)
    create_template(template, "Zuordnungsdatensatzliste")
    con = duckdb.connect(str(db_path))
    try:
        con.execute("update dq_export_gate_ru set gate_status='BLOCKED'")
    finally:
        con.close()
    with pytest.raises(RuntimeError, match="Export ist gesperrt"):
        export_module.build_nutzungsmeldung_xlsx(db_path, ["RU GmbH"], "RU", date(2026, 6, 1), date(2026, 6, 1), template)
