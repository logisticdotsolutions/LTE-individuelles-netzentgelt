from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import zuordnungen_export_module as module  # noqa: E402


class DummyConnection:
    def __init__(self) -> None:
        self.closed = False

    def close(self) -> None:
        self.closed = True


def _sample_rows() -> list[dict[str, object]]:
    return [
        {
            "locomotive_no": "91801234567-8",
            "usage_start": datetime(2026, 6, 9, 8, 15),
            "usage_end": datetime(2026, 6, 9, 12, 45),
            "performing_ru": "LTE DE - LTE Germany GmbH",
            "movement_count": 2,
            "user_vens": "1900100300001",
            "holder_market_partner_id": "1900100300002",
        }
    ]


def test_load_zuordnungen_workbook_hardens_layout_with_versioned_fallback(tmp_path: Path) -> None:
    workbook = module._load_zuordnungen_workbook(
        tmp_path / "Vorlage_Zuordnungen.xlsx"
    )
    worksheet = workbook["Zuordnungsdatensatzliste"]

    assert worksheet["A1"].value == "Zuordnung"
    assert worksheet["B2"].value == "Z01"
    assert tuple(
        worksheet.cell(row=6, column=column_number).value
        for column_number in range(1, 6)
    ) == module.ZUORDNUNGEN_HEADERS
    assert worksheet.max_column == 5


def test_build_zuordnungen_xlsx_writes_official_schema_and_values(
    monkeypatch,
    tmp_path: Path,
) -> None:
    db_path = tmp_path / "netzentgelt.duckdb"
    db_path.touch()
    dummy_connection = DummyConnection()

    monkeypatch.setattr(
        module.duckdb,
        "connect",
        lambda *_args, **_kwargs: dummy_connection,
    )
    monkeypatch.setattr(module, "_fetch_usage_segments", lambda **_kwargs: _sample_rows())
    monkeypatch.setattr(
        module,
        "_resolve_export_header",
        lambda **_kwargs: ("1900100300001", "LTE Germany GmbH"),
    )

    result = module.build_zuordnungen_xlsx(
        db_path=db_path,
        performing_ru_values=("LTE DE - LTE Germany GmbH",),
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=tmp_path / "Vorlage_Zuordnungen.xlsx",
    )

    assert dummy_connection.closed is True
    assert result.file_name == "Zuordnungen_LTE_DE_2026-06-09_bis_2026-06-09.xlsx"
    assert result.row_count == 1
    assert result.missing_required_field_count == 0

    workbook = load_workbook(BytesIO(result.content))
    worksheet = workbook["Zuordnungsdatensatzliste"]

    assert worksheet["A1"].value == "Zuordnung"
    assert worksheet["B2"].value == "Z01"
    assert worksheet["B3"].value == "1900100300001"
    assert worksheet["B4"].value == "LTE Germany GmbH"
    assert tuple(
        worksheet.cell(row=6, column=column_number).value
        for column_number in range(1, 6)
    ) == module.ZUORDNUNGEN_HEADERS
    assert worksheet.max_column == 5
    assert worksheet["A7"].value == "91801234567-8"
    assert worksheet["B7"].value == datetime(2026, 6, 9, 8, 15)
    assert worksheet["C7"].value == datetime(2026, 6, 9, 12, 45)
    assert worksheet["D7"].value == "1900100300001"
    assert worksheet["E7"].value == "1900100300002"
    assert worksheet["B7"].number_format == "dd.mm.yyyy hh:mm"
    assert worksheet["C7"].number_format == "dd.mm.yyyy hh:mm"


@pytest.mark.parametrize(
    "holding_market_partner_id",
    module.LTE_HOLDING_MARKET_PARTNER_IDS,
)
def test_build_zuordnungen_holding_xlsx_writes_fixed_holding_header(
    monkeypatch,
    tmp_path: Path,
    holding_market_partner_id: str,
) -> None:
    db_path = tmp_path / "netzentgelt.duckdb"
    db_path.touch()
    dummy_connection = DummyConnection()

    monkeypatch.setattr(
        module.duckdb,
        "connect",
        lambda *_args, **_kwargs: dummy_connection,
    )
    monkeypatch.setattr(
        module,
        "_fetch_holding_assignment_segments",
        lambda **_kwargs: _sample_rows(),
    )

    result = module.build_zuordnungen_holding_xlsx(
        db_path=db_path,
        holding_market_partner_id=holding_market_partner_id,
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=tmp_path / "Vorlage_Zuordnungen.xlsx",
    )

    assert dummy_connection.closed is True
    assert result.file_name == (
        f"Zuordnungen_LTE_Holding_{holding_market_partner_id}_"
        "2026-06-09_bis_2026-06-09.xlsx"
    )
    assert result.row_count == 1

    workbook = load_workbook(BytesIO(result.content))
    worksheet = workbook["Zuordnungsdatensatzliste"]

    assert worksheet["B3"].value == holding_market_partner_id
    assert worksheet["B4"].value == module.LTE_HOLDING_MARKET_PARTNER_NAME
    assert worksheet["A7"].value == "91801234567-8"
    assert worksheet["D7"].value == "1900100300001"


def test_build_zuordnungen_holding_xlsx_rejects_unknown_header_id(tmp_path: Path) -> None:
    db_path = tmp_path / "netzentgelt.duckdb"
    db_path.touch()

    with pytest.raises(ValueError, match="Unbekannte LTE-Holding-Marktpartner-ID"):
        module.build_zuordnungen_holding_xlsx(
            db_path=db_path,
            holding_market_partner_id="1900100300999",
            date_from=date(2026, 6, 9),
            date_to=date(2026, 6, 9),
        )
