from __future__ import annotations

"""AV01-Exporttests – Aufenthaltsabschnitte gegen UKL-Vorlage AV01.

Prüft:
- Tabellenblatt-Name 'Aufenthaltsabschnitt' (Vorlage-Kontrakt)
- Daten beginnen ab Zeile 8
- Netzstatus 'einfahrend' für clean_dir='E'
- Netzstatus 'netzintern' für IN_REPORT ohne Grenzrichtung
"""

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import duckdb
import pytest
from openpyxl import load_workbook

import sys
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import av01_export_module


def _create_minimal_av01_db(path: Path) -> None:
    """Minimale DuckDB-Fixture für AV01-Tests."""
    con = duckdb.connect(str(path))
    try:
        con.execute("""
            create table dq_export_gate_ru
            (loco_no varchar, performing_ru varchar, coverage_date date, gate_status varchar)
        """)
        con.execute("""
            insert into dq_export_gate_ru
            values ('91800000001-1','LTE DE GmbH','2026-06-09','READY')
        """)
        con.execute("""
            create table dq_global_export_blockers
            (blocker_date date, rule_id varchar, gate_status varchar)
        """)
        con.execute("""
            create table cfg_market_partner_mapping_effective (
                role_code varchar, source_value_normalized varchar, source_value varchar,
                official_company_name varchar, market_partner_id varchar,
                match_method varchar, match_score double
            )
        """)
        con.execute("""
            create table cfg_market_partner_role_effective (
                role_code varchar, company_name_normalized varchar,
                company_name_official varchar, market_partner_id varchar
            )
        """)
        con.execute("""
            create table core_loco_timeline (
                row_type varchar, loco_no varchar, performing_ru varchar,
                faulty_dir varchar, clean_dir varchar, report_scope varchar,
                sequence_ts timestamp, actual_departure_ts timestamp,
                actual_arrival_ts timestamp, origin_name varchar,
                destination_name varchar, needs_manual_review boolean
            )
        """)
        # Einfahrt: clean_dir='E' → netzstatus='einfahrend'
        con.execute("""
            insert into core_loco_timeline values
            ('MOVEMENT','91800000001-1','LTE DE GmbH',
             null,'E','IN_REPORT',
             '2026-06-09 08:00:00','2026-06-09 08:00:00','2026-06-09 09:00:00',
             'Grenzübergang Passau','Regensburg',false)
        """)
        # Netzintern: kein faulty_dir/clean_dir, IN_REPORT → netzstatus='netzintern'
        con.execute("""
            insert into core_loco_timeline values
            ('MOVEMENT','91800000001-1','LTE DE GmbH',
             null,null,'IN_REPORT',
             '2026-06-09 09:00:00','2026-06-09 09:00:00','2026-06-09 10:00:00',
             'Regensburg','München',false)
        """)
    finally:
        con.close()


def _blank_template(path: Path, sheet_name: str) -> None:
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = sheet_name
    wb.save(path)


@pytest.mark.integration
def test_av01_template_sheet_name_is_aufenthaltsabschnitt() -> None:
    """Vorlage AV01 muss das Blatt 'Aufenthaltsabschnitt' enthalten."""
    template_path = ROOT / "data" / "05_templates" / "Vorlage_Aufenthaltsabschnitt.xlsx"
    if not template_path.exists():
        pytest.skip("AV01-Vorlage nicht vorhanden")
    wb = load_workbook(template_path, read_only=True)
    try:
        assert "Aufenthaltsabschnitt" in wb.sheetnames, (
            f"Erwartetes Blatt 'Aufenthaltsabschnitt' fehlt. Gefunden: {wb.sheetnames}"
        )
    finally:
        wb.close()


@pytest.mark.integration
def test_av01_data_starts_at_row_8(tmp_path: Path) -> None:
    """AV01-Daten müssen ab Zeile 8 beginnen (UKL-Vorlage AV01 Kontrakt)."""
    db_path = tmp_path / "av01.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsabschnitt.xlsx"
    _create_minimal_av01_db(db_path)
    _blank_template(template, "Aufenthaltsabschnitt")

    result = av01_export_module.build_av01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    assert result.row_count == 2
    ws = load_workbook(BytesIO(result.content))["Aufenthaltsabschnitt"]
    assert ws.cell(row=7, column=1).value is None, "Zeile 7 darf keine Daten enthalten"
    assert ws.cell(row=8, column=1).value == "91800000001-1", "Erste Datenzeile ist Zeile 8"


@pytest.mark.integration
def test_av01_einfahrend_netzstatus(tmp_path: Path) -> None:
    """clean_dir='E' muss Netzstatus 'einfahrend' ergeben (Spalte 7)."""
    db_path = tmp_path / "av01.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsabschnitt.xlsx"
    _create_minimal_av01_db(db_path)
    _blank_template(template, "Aufenthaltsabschnitt")

    result = av01_export_module.build_av01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    ws = load_workbook(BytesIO(result.content))["Aufenthaltsabschnitt"]
    netzstatus_first_row = ws.cell(row=8, column=7).value
    assert netzstatus_first_row == "einfahrend", (
        f"Erste Zeile (clean_dir='E') muss 'einfahrend' sein, war: {netzstatus_first_row}"
    )


@pytest.mark.integration
def test_av01_netzintern_netzstatus(tmp_path: Path) -> None:
    """Bewegung ohne Grenzrichtung (IN_REPORT) muss Netzstatus 'netzintern' ergeben."""
    db_path = tmp_path / "av01.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsabschnitt.xlsx"
    _create_minimal_av01_db(db_path)
    _blank_template(template, "Aufenthaltsabschnitt")

    result = av01_export_module.build_av01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    ws = load_workbook(BytesIO(result.content))["Aufenthaltsabschnitt"]
    netzstatus_second_row = ws.cell(row=9, column=7).value
    assert netzstatus_second_row == "netzintern", (
        f"Zweite Zeile (kein dir, IN_REPORT) muss 'netzintern' sein, war: {netzstatus_second_row}"
    )


@pytest.mark.integration
def test_av01_departure_and_arrival_columns(tmp_path: Path) -> None:
    """AV01 Spalte 3 (Beginn) und Spalte 5 (Ende) müssen datetime-Werte enthalten."""
    db_path = tmp_path / "av01.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsabschnitt.xlsx"
    _create_minimal_av01_db(db_path)
    _blank_template(template, "Aufenthaltsabschnitt")

    result = av01_export_module.build_av01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    ws = load_workbook(BytesIO(result.content))["Aufenthaltsabschnitt"]
    beginn = ws.cell(row=8, column=3).value
    ende = ws.cell(row=8, column=5).value
    assert beginn is not None, "Beginn (Spalte 3) darf nicht leer sein"
    assert isinstance(beginn, datetime), f"Beginn muss datetime sein, war: {type(beginn)}"
    assert ende is not None, "Ende (Spalte 5) darf nicht leer sein"
    assert isinstance(ende, datetime), f"Ende muss datetime sein, war: {type(ende)}"
