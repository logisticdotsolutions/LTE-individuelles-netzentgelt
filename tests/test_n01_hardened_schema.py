from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import n01_hardened_export_module as module  # noqa: E402


class DummyConnection:
    def close(self) -> None:
        pass


def test_n01_schema_has_exactly_five_columns(monkeypatch, tmp_path: Path) -> None:
    db_path = tmp_path / "netzentgelt.duckdb"
    db_path.touch()
    row = {
        "locomotive_no": "91801234567-8",
        "usage_start": datetime(2026, 6, 9, 8, 15),
        "usage_end": datetime(2026, 6, 9, 12, 45),
        "performing_ru": "LTE DE - LTE Germany GmbH",
        "movement_count": 2,
        "user_vens": "1900100300001",
        "holder_market_partner_id": "1900100300393",
    }
    monkeypatch.setattr(module.duckdb, "connect", lambda *_args, **_kwargs: DummyConnection())
    monkeypatch.setattr(module, "_fetch_usage_segments", lambda **_kwargs: [row])
    monkeypatch.setattr(module, "_resolve_export_header", lambda **_kwargs: ("1900100300001", "LTE Germany GmbH"))

    result = module.build_hardened_n01_xlsx(
        db_path=db_path,
        performing_ru_values=("LTE DE - LTE Germany GmbH",),
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
    )

    worksheet = load_workbook(BytesIO(result.content))["Zuordnungsdatensatzliste"]
    assert worksheet["B2"].value == "N01"
    assert worksheet.max_column == 5
    assert tuple(worksheet.cell(row=6, column=column).value for column in range(1, 6)) == module.N01_HEADERS
    assert worksheet["E7"].value == "1900100300393"
