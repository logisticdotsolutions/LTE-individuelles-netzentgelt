from __future__ import annotations

"""Spaltenvertrag-Tests für XLSX-Exporte gegen die UKL-Vorlagen.

Prüft:
- Aufenthaltsereignis Spalte 4 (Zeitpunkt/event_ts) – bisher fehlende Assertion
- Template-Blattnamen-Kontrakt (Vorlage muss erwartetes Blatt enthalten)
- Nutzungsmeldung data_start_row-Kontrakt (Daten ab Zeile 7)
- Aufenthaltsereignis data_start_row-Kontrakt (Daten ab Zeile 8)
"""

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import sys

import duckdb
import pytest
from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import export_module  # noqa: E402


# ---------------------------------------------------------------------------
# Minimal-Fixture-Helfer
# ---------------------------------------------------------------------------

def _create_minimal_export_db(path: Path) -> None:
    """Minimale DuckDB-Fixture mit einem DE-relevanten Einfahrts-Bewegungsdatensatz."""
    con = duckdb.connect(str(path))
    try:
        con.execute("""
            create table core_usage_assignment_segments (
                usage_segment_id varchar, loco_no varchar, tfze_or_tens varchar,
                segment_start_utc timestamp, segment_end_utc timestamp,
                performing_ru varchar, movement_count bigint, user_vens varchar,
                holder_market_partner_id varchar, holder_name varchar,
                export_blocking_movement_rows bigint
            )
        """)
        con.execute("""
            insert into core_usage_assignment_segments
            values ('SEG-1','91800000001-1','91800000001-1',
                    '2026-06-09 08:00:00','2026-06-09 09:00:00',
                    'LTE DE GmbH',1,'1900100300001','9000000000001','Holder AG',0)
        """)
        con.execute("""
            create table core_usage_assignment_segment_movements
            (usage_segment_id varchar, actual_departure_ts timestamp)
        """)
        con.execute("""
            insert into core_usage_assignment_segment_movements
            values ('SEG-1','2026-06-09 08:00:00')
        """)
        con.execute("""
            create table dq_export_gate_ru
            (loco_no varchar, performing_ru varchar, coverage_date date, gate_status varchar)
        """)
        con.execute("""
            insert into dq_export_gate_ru
            values ('91800000001-1','LTE DE GmbH','2026-06-09','READY')
        """)
        con.execute("create table dq_global_export_blockers (blocker_date date, rule_id varchar, gate_status varchar)")
        con.execute("""
            create table cfg_market_partner_mapping_effective (
                role_code varchar, source_value_normalized varchar, source_value varchar,
                official_company_name varchar, market_partner_id varchar,
                match_method varchar, match_score double
            )
        """)
        con.execute("""
            create table cfg_market_partner_role_effective (
                role_code varchar, company_name_normalized varchar,
                company_name_official varchar, market_partner_id varchar
            )
        """)
        con.execute("""
            create table core_loco_timeline (
                row_type varchar, loco_no varchar, performing_ru varchar,
                faulty_dir varchar, clean_dir varchar, report_scope varchar,
                sequence_ts timestamp, actual_departure_ts timestamp,
                actual_arrival_ts timestamp, origin_name varchar,
                destination_name varchar, needs_manual_review boolean
            )
        """)
        # Einfahrt: clean_dir='E' → event_ts = actual_departure_ts, event_location = origin_name
        con.execute("""
            insert into core_loco_timeline
            values ('MOVEMENT','91800000001-1','LTE DE GmbH',
                    null,'E','IN_REPORT',
                    '2026-06-09 08:00:00','2026-06-09 08:00:00','2026-06-09 09:00:00',
                    'Grenzübergang Passau','Regensburg',false)
        """)
    finally:
        con.close()


def _blank_template(path: Path, sheet_name: str) -> None:
    wb = Workbook()
    wb.active.title = sheet_name
    wb.save(path)


# ---------------------------------------------------------------------------
# Aufenthaltsereignis – Spalte 4 (Zeitpunkt / event_ts)
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_aufenthaltsereignis_column_4_is_event_timestamp(tmp_path: Path) -> None:
    """Spalte 4 (Zeitpunkt) muss den Zeitstempel des Aufenthaltsereignisses enthalten.

    Diese Assertion fehlte im Basis-Integrationstest. Der Zeitstempel ist ein
    Pflichtfeld der Vorlage AE01 und muss korrekt befüllt sein.
    """
    db_path = tmp_path / "ae01.duckdb"
    template = tmp_path / "Vorlage_Aufenthaltsereignis.xlsx"
    _create_minimal_export_db(db_path)
    _blank_template(template, "Aufenthaltsereignisse")

    result = export_module.build_aufenthaltsereignis_xlsx(
        db_path, ["LTE DE GmbH"], "LTE_DE",
        date(2026, 6, 9), date(2026, 6, 9), template,
    )

    assert result.row_count == 1
    assert result.missing_required_field_count == 0

    ws = load_workbook(BytesIO(result.content))["Aufenthaltsereignisse"]
    first_data_row = 8

    # Spalte 1: TfzE oder tEns
    assert ws.cell(row=first_data_row, column=1).value == "91800000001-1"
    # Spalte 2: vEns (Performing RU)
    assert ws.cell(row=first_data_row, column=2).value == "LTE DE GmbH"
    # Spalte 3: Ort (Border Point / Bewegungsort)
    assert ws.cell(row=first_data_row, column=3).value == "Grenzübergang Passau"
    # Spalte 4: Zeitpunkt – explizit prüfen (war bisher nicht abgedeckt)
    zeitpunkt = ws.cell(row=first_data_row, column=4).value
    assert zeitpunkt is not None, "Zeitpunkt (Spalte 4) darf nicht leer sein"
    assert isinstance(zeitpunkt, datetime), f"Zeitpunkt muss ein datetime sein, war: {type(zeitpunkt)}"
    # Spalte 5: Netzstatus
    assert ws.cell(row=first_data_row, column=5).value == "einfahrend"


# ---------------------------------------------------------------------------
# Nutzungsmeldung – data_start_row und Pflichtfelder
# ---------------------------------------------------------------------------

@pytest.mark.integration
def test_nutzungsmeldung_data_starts_at_row_7(tmp_path: Path) -> None:
    """Nutzungsmeldung-Daten müssen ab Zeile 7 beginnen (UKL-Vorlage N01 Kontrakt)."""
    db_path = tmp_path / "n01.duckdb"
    template = tmp_path / "Vorlage_Nutzungsmeldung.xlsx"
    _create_minimal_export_db(db_path)
    _blank_template(template, "Zuordnungsdatensatzliste")

    result = export_module.build_nutzungsmeldung_xlsx(
        db_path, ["LTE DE GmbH"], "LTE_DE",
        date(2026, 6, 9), date(2026, 6, 9), template,
    )
    assert result.row_count == 1
    ws = load_workbook(BytesIO(result.content))["Zuordnungsdatensatzliste"]

    # Zeile 6 muss leer sein (Kopfzeile der Vorlage, keine Daten)
    assert ws.cell(row=6, column=1).value is None, "Zeile 6 darf keine Datenzelle sein"
    # Daten beginnen ab Zeile 7
    assert ws.cell(row=7, column=1).value == "91800000001-1", "Erste Datenzeile ist Zeile 7"
    # Spalte 6 (Übernahmeanfrage oder Übergabemeldung?) muss leer bleiben
    assert ws.cell(row=7, column=6).value is None, "Spalte 6 (Übergabemeldung) muss leer sein"


# ---------------------------------------------------------------------------
# Template-Blattnamen-Kontrakt
# ---------------------------------------------------------------------------

TEMPLATE_SHEET_CONTRACTS = [
    ("Vorlage_Nutzungsmeldung.xlsx", "Zuordnungsdatensatzliste"),
    ("Vorlage_Aufenthaltsereignis.xlsx", "Aufenthaltsereignisse"),
    ("Vorlage_Zuordnungen.xlsx", "Zuordnungsdatensatzliste"),
]


@pytest.mark.integration
@pytest.mark.parametrize("template_file,expected_sheet", TEMPLATE_SHEET_CONTRACTS)
def test_xlsx_template_contains_expected_sheet(template_file: str, expected_sheet: str) -> None:
    """Jede UKL-XLSX-Vorlage muss das vom Export erwartete Tabellenblatt enthalten.

    Schlägt ein Blattnamen-Kontrakt fehl, schlägt auch der Export zur Laufzeit fehl.
    Dieser Test pinnt den Vertrag, damit Template-Änderungen sofort sichtbar werden.
    """
    template_path = ROOT / "data" / "05_templates" / template_file
    if not template_path.exists():
        pytest.skip(f"Template nicht vorhanden: {template_path}")

    wb = load_workbook(template_path, read_only=True)
    try:
        assert expected_sheet in wb.sheetnames, (
            f"{template_file}: Erwartetes Blatt '{expected_sheet}' nicht gefunden. "
            f"Vorhandene Blätter: {wb.sheetnames}"
        )
    finally:
        wb.close()
