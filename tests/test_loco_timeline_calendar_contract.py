from __future__ import annotations

from datetime import date
from io import BytesIO
from pathlib import Path
import sys

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from loco_timeline_calendar_runtime_module import (  # noqa: E402
    build_loco_timeline_day_summary,
    build_loco_timeline_segments,
    build_loco_timeline_xlsx,
    classify_timeline_status,
    filter_loco_timeline_segments,
)
from loco_timeline_review_queue_module import build_loco_timeline_review_queue  # noqa: E402


def test_status_priority_prefers_manual_check_before_gap_or_overlap():
    assert (
        classify_timeline_status(
            row_type="GAP",
            is_de_relevant=True,
            holder="LTE Holding",
            performing_ru="LTE DE",
            rules="R010",
            message="Relevante Lücke muss geprüft werden",
            decision_reason="",
        )
        == "Prüfen"
    )
    assert (
        classify_timeline_status(
            row_type="MOVEMENT",
            is_de_relevant=True,
            holder="LTE Holding",
            performing_ru="LTE DE",
            rules="",
            message="",
            decision_reason="Overlap bis 5 Minuten ignorieren",
        )
        == "Overlap"
    )


def test_segments_are_split_by_day_and_include_context_days():
    source = pd.DataFrame(
        [
            {
                "loco_no": "193 001",
                "holder_name": "LTE Holding",
                "performing_ru": "LTE DE",
                "row_type": "MOVEMENT",
                "report_scope": "IN_REPORT",
                "period_start_utc": "2026-06-10T23:30:00Z",
                "period_end_utc": "2026-06-11T00:30:00Z",
                "cal_route_type_home": "Inland",
            },
            {
                "loco_no": "193 002",
                "holder_name": "LTE Holding",
                "performing_ru": "LTE NL",
                "row_type": "MOVEMENT",
                "report_scope": "IN_REPORT",
                "period_start_utc": "2026-06-12T08:00:00Z",
                "period_end_utc": "2026-06-12T09:00:00Z",
                "cal_route_type_home": "Inland",
            },
        ]
    )

    segments = build_loco_timeline_segments(
        source,
        date_from=date(2026, 6, 11),
        date_to=date(2026, 6, 11),
        context_days=1,
    )

    assert set(segments["Meldetag"]) == {"2026-06-10", "2026-06-11", "2026-06-12"}
    assert segments.loc[segments["Meldetag"].eq("2026-06-10"), "Im Filterzeitraum"].eq(False).all()
    assert segments.loc[segments["Meldetag"].eq("2026-06-11"), "Im Filterzeitraum"].eq(True).all()


def test_filter_and_summary_keep_highest_status_per_loco_day():
    source = pd.DataFrame(
        [
            {
                "loco_no": "193 001",
                "holder_name": "LTE Holding",
                "performing_ru": "LTE DE",
                "row_type": "MOVEMENT",
                "report_scope": "IN_REPORT",
                "period_start_utc": "2026-06-11T08:00:00Z",
                "period_end_utc": "2026-06-11T09:00:00Z",
            },
            {
                "loco_no": "193 001",
                "holder_name": "LTE Holding",
                "performing_ru": "LTE DE",
                "row_type": "GAP",
                "report_scope": "IN_REPORT",
                "period_start_utc": "2026-06-11T09:00:00Z",
                "period_end_utc": "2026-06-11T10:00:00Z",
            },
        ]
    )

    segments = build_loco_timeline_segments(
        source,
        date_from=date(2026, 6, 11),
        date_to=date(2026, 6, 11),
        context_days=0,
    )
    filtered = filter_loco_timeline_segments(segments, only_problem_cases=True)
    summary = build_loco_timeline_day_summary(segments)

    assert len(filtered) == 1
    assert filtered.iloc[0]["Status"] == "GAP"
    assert summary.iloc[0]["Status"] == "GAP"
    assert int(summary.iloc[0]["Problemsegmente"]) == 1


def test_loco_timeline_xlsx_contains_review_sheets():
    source = pd.DataFrame(
        [
            {
                "loco_no": "193 001",
                "holder_name": "LTE Holding",
                "performing_ru": "LTE DE",
                "row_type": "MOVEMENT",
                "report_scope": "IN_REPORT",
                "period_start_utc": "2026-06-11T08:00:00Z",
                "period_end_utc": "2026-06-11T09:00:00Z",
            }
        ]
    )
    segments = build_loco_timeline_segments(
        source,
        date_from=date(2026, 6, 11),
        date_to=date(2026, 6, 11),
        context_days=0,
    )
    summary = build_loco_timeline_day_summary(segments)

    payload = build_loco_timeline_xlsx(segments, summary)
    workbook = load_workbook(BytesIO(payload), read_only=True)

    assert workbook.sheetnames == ["Tagesstatus", "Segmente", "Legende"]
    assert workbook["Tagesstatus"].max_row == 2
    assert workbook["Segmente"].max_row == 2
    assert workbook["Legende"].max_row >= 7


def test_review_queue_orders_problem_loco_days_by_priority():
    segments = pd.DataFrame(
        [
            {
                "Meldetag": "2026-06-11",
                "Loknummer": "193 002",
                "Halter": "LTE Holding",
                "Nutzer / PerformingRU": "LTE NL",
                "Status": "GAP",
                "StatusPriorität": 30,
                "StartMinute": 480,
                "EndMinute": 540,
                "Regeln": "",
                "Meldung": "",
                "Begründung": "",
            },
            {
                "Meldetag": "2026-06-11",
                "Loknummer": "193 001",
                "Halter": "LTE Holding",
                "Nutzer / PerformingRU": "LTE DE",
                "Status": "Prüfen",
                "StatusPriorität": 50,
                "StartMinute": 600,
                "EndMinute": 660,
                "Regeln": "R010",
                "Meldung": "Prüfen",
                "Begründung": "Manuelle Entscheidung erforderlich",
            },
            {
                "Meldetag": "2026-06-11",
                "Loknummer": "193 003",
                "Halter": "LTE Holding",
                "Nutzer / PerformingRU": "LTE DE",
                "Status": "Zugewiesen",
                "StatusPriorität": 20,
                "StartMinute": 700,
                "EndMinute": 760,
                "Regeln": "",
                "Meldung": "",
                "Begründung": "",
            },
        ]
    )

    queue = build_loco_timeline_review_queue(segments)

    assert queue["Loknummer"].tolist() == ["193 001", "193 002"]
    assert queue.iloc[0]["Status"] == "Prüfen"
    assert queue.iloc[0]["Erste Uhrzeit"] == "10:00"
    assert queue.iloc[1]["Status"] == "GAP"
