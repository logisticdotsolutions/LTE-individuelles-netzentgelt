from __future__ import annotations

"""AB01-Exporttests – Abstellungen gegen UKL-Vorlage AB01.

Prüft:
- Tabellenblatt-Name 'Abstellungen' (Vorlage-Kontrakt)
- Daten beginnen ab Zeile 8
- Art-Spalte enthält 'TfzE nicht in Nutzung' (offene UKL-Frage Spec 23.7)
- Keine Abstellungen → row_count=0, leere aber stabile XLSX-Ausgabe
- Pflichtfelder Beginn und Ende sind datetime-Werte
"""

from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import duckdb
import pytest
from openpyxl import load_workbook, Workbook

import sys
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

import ab01_export_module


def _create_minimal_ab01_db(path: Path, *, include_stand: bool = True) -> None:
    """Minimale DuckDB-Fixture für AB01-Tests."""
    con = duckdb.connect(str(path))
    try:
        con.execute("""
            create table dq_export_gate_ru
            (loco_no varchar, performing_ru varchar, coverage_date date, gate_status varchar)
        """)
        con.execute("""
            insert into dq_export_gate_ru
            values ('91800000001-1','LTE DE GmbH','2026-06-09','READY')
        """)
        con.execute("""
            create table dq_global_export_blockers
            (blocker_date date, rule_id varchar, gate_status varchar)
        """)
        con.execute("""
            create table cfg_market_partner_mapping_effective (
                role_code varchar, source_value_normalized varchar, source_value varchar,
                official_company_name varchar, market_partner_id varchar,
                match_method varchar, match_score double
            )
        """)
        con.execute("""
            create table cfg_market_partner_role_effective (
                role_code varchar, company_name_normalized varchar,
                company_name_official varchar, market_partner_id varchar
            )
        """)
        con.execute("""
            create table core_loco_stand_candidates (
                run_id varchar, loco_no varchar,
                transport_number varchar, next_transport_number varchar,
                location_name varchar,
                stand_from_utc timestamp, stand_to_utc timestamp,
                stand_duration_minutes double,
                performing_ru varchar,
                report_scope varchar, next_report_scope varchar,
                stand_class varchar, suggested_action varchar,
                source_table varchar, source_row_id bigint
            )
        """)
        if include_stand:
            con.execute("""
                insert into core_loco_stand_candidates values
                ('RUN-1','91800000001-1',
                 'T-100','T-200',
                 'Hamburg Hbf',
                 '2026-06-09 08:00:00','2026-06-09 18:00:00',
                 600.0,
                 'LTE DE GmbH',
                 'IN_REPORT','IN_REPORT',
                 'POTENTIAL_COLD_STAND',
                 'Standzeit ueber 8 Stunden am selben Ort.',
                 'core_loco_timeline',1)
            """)
    finally:
        con.close()


def _blank_template(path: Path, sheet_name: str) -> None:
    wb = Workbook()
    wb.active.title = sheet_name
    wb.save(path)


@pytest.mark.integration
def test_ab01_template_sheet_name_is_abstellungen() -> None:
    """Vorlage AB01 muss das Blatt 'Abstellungen' enthalten."""
    template_path = ROOT / "data" / "05_templates" / "Vorlage_Abstellungen.xlsx"
    if not template_path.exists():
        pytest.skip("AB01-Vorlage nicht vorhanden")
    wb = load_workbook(template_path, read_only=True)
    try:
        assert "Abstellungen" in wb.sheetnames, (
            f"Erwartetes Blatt 'Abstellungen' fehlt. Gefunden: {wb.sheetnames}"
        )
    finally:
        wb.close()


@pytest.mark.integration
def test_ab01_data_starts_at_row_8(tmp_path: Path) -> None:
    """AB01-Daten müssen ab Zeile 8 beginnen (UKL-Vorlage AB01 Kontrakt)."""
    db_path = tmp_path / "ab01.duckdb"
    template = tmp_path / "Vorlage_Abstellungen.xlsx"
    _create_minimal_ab01_db(db_path, include_stand=True)
    _blank_template(template, "Abstellungen")

    result = ab01_export_module.build_ab01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    assert result.row_count == 1
    ws = load_workbook(BytesIO(result.content))["Abstellungen"]
    assert ws.cell(row=7, column=1).value is None, "Zeile 7 darf keine Daten enthalten"
    assert ws.cell(row=8, column=1).value == "91800000001-1", "Erste Datenzeile ist Zeile 8"


@pytest.mark.integration
def test_ab01_art_is_tfze_nicht_in_nutzung(tmp_path: Path) -> None:
    """Art-Spalte (Spalte 3) muss 'TfzE nicht in Nutzung' enthalten (Spec 23.7 offen).

    Bis zur schriftlichen Klärung durch UKL, wie Kaltabstellungen formal exportiert
    werden, ist 'TfzE nicht in Nutzung' der einzige zulässige Vorlage-Wert für
    POTENTIAL_COLD_STAND-Einträge.
    """
    db_path = tmp_path / "ab01.duckdb"
    template = tmp_path / "Vorlage_Abstellungen.xlsx"
    _create_minimal_ab01_db(db_path, include_stand=True)
    _blank_template(template, "Abstellungen")

    result = ab01_export_module.build_ab01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    ws = load_workbook(BytesIO(result.content))["Abstellungen"]
    art = ws.cell(row=8, column=3).value
    assert art == "TfzE nicht in Nutzung", (
        f"Art muss 'TfzE nicht in Nutzung' sein, war: {art!r}"
    )


@pytest.mark.integration
def test_ab01_empty_stands_produces_stable_output(tmp_path: Path) -> None:
    """Keine Abstellungen im Zeitraum → row_count=0, stabile leere XLSX-Ausgabe."""
    db_path = tmp_path / "ab01_empty.duckdb"
    template = tmp_path / "Vorlage_Abstellungen.xlsx"
    _create_minimal_ab01_db(db_path, include_stand=False)
    _blank_template(template, "Abstellungen")

    result = ab01_export_module.build_ab01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    assert result.row_count == 0
    assert len(result.content) > 0, "Leere XLSX darf nicht leer sein (Vorlage muss erhalten bleiben)"
    ws = load_workbook(BytesIO(result.content))["Abstellungen"]
    assert ws.cell(row=8, column=1).value is None, "Zeile 8 muss bei leerer Ausgabe leer sein"


@pytest.mark.integration
def test_ab01_beginn_and_ende_are_datetime(tmp_path: Path) -> None:
    """AB01 Beginn (Spalte 4) und Ende (Spalte 5) müssen datetime-Werte enthalten."""
    db_path = tmp_path / "ab01.duckdb"
    template = tmp_path / "Vorlage_Abstellungen.xlsx"
    _create_minimal_ab01_db(db_path, include_stand=True)
    _blank_template(template, "Abstellungen")

    result = ab01_export_module.build_ab01_xlsx(
        db_path=db_path,
        performing_ru_values=["LTE DE GmbH"],
        export_label="LTE_DE",
        date_from=date(2026, 6, 9),
        date_to=date(2026, 6, 9),
        template_path=template,
    )

    ws = load_workbook(BytesIO(result.content))["Abstellungen"]
    beginn = ws.cell(row=8, column=4).value
    ende = ws.cell(row=8, column=5).value
    assert beginn is not None, "Beginn (Spalte 4) darf nicht leer sein"
    assert isinstance(beginn, datetime), f"Beginn muss datetime sein, war: {type(beginn)}"
    assert ende is not None, "Ende (Spalte 5) darf nicht leer sein"
    assert isinstance(ende, datetime), f"Ende muss datetime sein, war: {type(ende)}"
